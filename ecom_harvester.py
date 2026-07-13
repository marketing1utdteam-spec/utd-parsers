#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════╗
║            Shopify Store Contact Harvester  v1.0                   ║
║   Live Shopify-Merchant & Deliverable-Contact Collector            ║
║                                                                    ║
║   Third parser in the family (b2b_harvester → influencer_scraper   ║
║   → THIS). Same engine, different prey:                            ║
║     • targets ACTIVE ecommerce STORES running on Shopify           ║
║       (not agencies, not bloggers) — to pitch UTD themes           ║
║     • hard Shopify gate: page must contain cdn.shopify.com /       ║
║       myshopify.com / window.Shopify / Shopify.theme               ║
║     • extracts the store's CURRENT THEME name from Shopify.theme   ║
║     • classifies store INDUSTRY by homepage keywords               ║
║     • email-domain MUST match the store's own domain               ║
║     • domain MUST have MX (DNS-over-HTTPS, no port 25 needed)      ║
║     • site MUST be live and NOT parked / for-sale / suspended      ║
║                                                                    ║
║   Reusable │ Resumable │ Rotating API keys │ Excel + Google Sheets ║
╚═══════════════════════════════════════════════════════════════════╝

Usage:
    python ecom_harvester.py                  # normal run
    python ecom_harvester.py --reset          # clear state, start fresh
    python ecom_harvester.py --key sk-ant-..  # add Claude key (optional)
    Ctrl+C                                    # graceful stop + save

To retarget, edit the PROFILE block below — nothing else needs to change.
"""

import os, re, sys, json, time, signal, logging, hashlib, argparse, random
from datetime import datetime
from urllib.parse import urlparse
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════════
#   🎯  PROFILE  —  edit THIS block to retarget the whole harvester
# ═══════════════════════════════════════════════════════════════════

PROFILE = {
    "name": "shopify_ecom_stores",

    # Product niches (drive query generation): <niche> × <intent> × <geo>.
    "niches": [
        "fashion", "clothing", "streetwear", "menswear", "womenswear",
        "jewelry", "accessories", "lingerie", "shoes", "sneakers",
        "bags", "watches", "swimwear",
        "beauty", "cosmetics", "skincare", "haircare", "fragrance",
        "home decor", "furniture", "candles", "kitchenware", "bedding",
        "food", "coffee", "tea", "chocolate", "snacks",
        "supplements", "vitamins", "wellness",
        "fitness", "sports", "activewear", "yoga", "outdoor", "camping",
        "pets", "pet supplies", "dog accessories",
        "electronics", "gadgets", "phone accessories", "smart home",
        "toys", "kids", "baby", "nursery",
        "art prints", "wall art", "stationery", "crafts", "handmade",
        "auto parts", "car accessories", "motorcycle gear",
    ],

    # Store industry classifier: label → homepage keyword markers.
    # Scored by hit count in <title> + meta description + visible text.
    "industries": {
        "Fashion & Apparel": [
            "fashion", "clothing", "apparel", "wear", "dress", "menswear",
            "womenswear", "streetwear", "t-shirt", "tee", "hoodie", "denim",
            "jeans", "lingerie", "swimwear", "outfit", "wardrobe", "shoes",
            "sneaker", "boots", "footwear",
        ],
        "Jewelry & Accessories": [
            "jewelry", "jewellery", "necklace", "bracelet", "earring",
            "ring", "pendant", "watch", "watches", "accessor", "handbag",
            "bags", "wallet", "sunglasses", "gold plated", "sterling silver",
        ],
        "Beauty & Cosmetics": [
            "beauty", "cosmetic", "skincare", "skin care", "makeup",
            "serum", "moisturizer", "cleanser", "fragrance", "perfume",
            "haircare", "hair care", "shampoo", "nail", "lipstick", "spf",
        ],
        "Sports & Fitness": [
            "fitness", "gym", "workout", "training", "sportswear",
            "activewear", "athletic", "yoga", "running", "cycling",
            "outdoor", "camping", "hiking", "sports", "athlete",
        ],
        "Food & Beverage": [
            "coffee", "tea", "chocolate", "snack", "food", "beverage",
            "drink", "wine", "brewery", "sauce", "spice", "gourmet",
            "organic food", "roasted", "recipe", "tasty", "flavor",
        ],
        "Electronics & Tech": [
            "electronics", "gadget", "tech", "headphone", "earbud",
            "speaker", "charger", "cable", "smart home", "drone",
            "camera", "keyboard", "phone case", "device",
        ],
        "Home & Furniture": [
            "home decor", "furniture", "sofa", "candle", "rug", "bedding",
            "kitchen", "tableware", "lighting", "lamp", "interior",
            "homeware", "vase", "cushion", "curtain", "living room",
        ],
        "Kids & Toys": [
            "toys", "kids", "baby", "children", "nursery", "plush",
            "montessori", "toddler", "playroom", "newborn",
        ],
        "Pets": [
            "pet", "dog", "cat", "puppy", "kitten", "pet supplies",
            "pet food", "leash", "grooming", "aquarium",
        ],
        "Health & Supplements": [
            "supplement", "vitamin", "protein", "wellness", "nutrition",
            "collagen", "probiotic", "immune", "cbd", "herbal", "remedy",
        ],
        "Art & Crafts": [
            "art print", "wall art", "poster", "canvas", "craft",
            "handmade", "stationery", "sticker", "embroidery", "ceramic",
            "pottery", "illustration", "artisan",
        ],
        "Auto & Moto": [
            "auto parts", "car accessories", "car care", "motorcycle",
            "moto", "detailing", "tires", "wheels", "4x4", "offroad",
        ],
    },
    "industry_default": "Other",

    # Words that CONFIRM it is a real consumer store (not a service site).
    "positive": [
        "add to cart", "shop now", "free shipping", "new arrivals",
        "best sellers", "shop all", "our products", "checkout",
        "sold out", "sale", "collections", "buy now", "gift card",
        "returns policy", "track your order", "worldwide shipping",
    ],

    # Words that DISQUALIFY: agencies, service providers, marketplaces.
    "negative": [
        "web design services", "we build stores", "we build shopify",
        "shopify agency", "shopify experts", "shopify partner",
        "development agency", "design agency", "digital agency",
        "marketing agency", "seo services", "our clients", "case studies",
        "hire us", "book a call", "free consultation", "request a quote",
        "theme documentation", "app for shopify", "become a vendor",
        "become a seller", "start selling", "marketplace platform",
        "sell on our marketplace", "dropshipping course", "we are hiring",
    ],

    # One-line question fed to the optional Claude validator.
    "claude_question": (
        "Is this website an ACTUAL online store (ecommerce shop selling "
        "products directly to customers), NOT an agency, app/theme vendor, "
        "marketplace, blog, or service company?"
    ),
}

# ═══════════════════════════════════════════════════════════════════
#   ⚙️  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

# Claude API key for AI validation of borderline cases (optional).
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Google CSE key / id pairs — rotate on quota/block.
_API_KEYS = [k.strip() for k in os.environ.get("GOOGLE_API_KEYS", "").split(",") if k.strip()]
_CSE_IDS  = [c.strip() for c in os.environ.get("GOOGLE_CSE_IDS", "").split(",") if c.strip()]
API_PAIRS = [{"api_key": k, "cse_id": _CSE_IDS[i % len(_CSE_IDS)]}
             for i, k in enumerate(_API_KEYS)]
# KEY_SLICE ("start:end") gives this parser a DEDICATED slice of the shared key
# pool so the three parsers never compete for the same daily CSE quota (the
# contention that caused the afternoon 429 storms). Falls back to the full pool.
_KEY_SLICE = os.environ.get("KEY_SLICE", "").strip()
if _KEY_SLICE and ":" in _KEY_SLICE:
    _a, _b = _KEY_SLICE.split(":")
    API_PAIRS = API_PAIRS[(int(_a) if _a else None):(int(_b) if _b else None)] or API_PAIRS

# ─── File paths (STATE_DIR keeps logs/state on the host volume) ──
_STATE_DIR    = os.environ.get("STATE_DIR", ".")
os.makedirs(_STATE_DIR, exist_ok=True)
OUTPUT_EXCEL  = os.path.join(_STATE_DIR, "ecom_stores_v1.xlsx")
STATE_FILE    = os.path.join(_STATE_DIR, "ecom_v1_state.json")
VISITED_FILE  = os.path.join(_STATE_DIR, "ecom_v1_visited.json")
LOG_FILE      = os.path.join(_STATE_DIR, "ecom_v1.log")

# ─── Google Sheets ───────────────────────────────────────────────
#  ECOM_SHEET_ID is REQUIRED (no default — the sheet is created manually
#  and its id is passed via env / GHA workflow).
SHEETS_SPREADSHEET_ID = os.environ.get("ECOM_SHEET_ID", "").strip()
SHEETS_WORKSHEET_NAME = os.environ.get("ECOM_SHEET_TAB", "Contacts")
SHEETS_CREDS_FILE     = os.environ.get("GOOGLE_CREDS_FILE", os.path.join(_STATE_DIR, "google_credentials.json"))

# ─── Session limits ─────────────────────────────────────────────
MAX_GOOGLE_QUERIES  = int(os.environ.get("MAX_GOOGLE_QUERIES", "50"))  # new unique Google searches per session
RESULTS_PER_QUERY   = 10    # Google results per query (max 10)
SEARCH_PAGES        = int(os.environ.get("SEARCH_PAGES", "2"))  # result pages per query (each page = 1 API call)
MAX_SUBPAGES        = 3     # extra subpages checked per domain
MIN_SCORE           = 60    # keep contacts scoring >= this (0-100)

# ─── Mailbox verification (optional free-tier API, as in b2b) ────
MAILBOX_PROVIDER = os.environ.get("MAILBOX_PROVIDER", "")  # "millionverifier" | "reoon" | "" (off)
MAILBOX_API_KEYS = [k.strip() for k in os.environ.get("MILLIONVERIFIER_KEYS", "").split(",") if k.strip()]
MAILBOX_DROP_CATCHALL = True   # drop catch-all domains (can't confirm mailbox)
MAILBOX_DROP_UNKNOWN  = True   # drop unverifiable results (max precision)

# ─── Delays (seconds) ───────────────────────────────────────────
DELAY_PAGE   = 1.2
DELAY_URL    = 1.6
DELAY_QUERY  = 2.4
API_ROT_WAIT = 4.0

# ─── Rotating user agents ───────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

# ═══════════════════════════════════════════════════════════════════
#   🛍️  SHOPIFY DETECTION  (the hard gate)
# ═══════════════════════════════════════════════════════════════════

#  A page qualifies as a Shopify storefront when ANY of these markers
#  is present in the raw HTML.
SHOPIFY_MARKERS = ("cdn.shopify.com", "myshopify.com", "window.Shopify", "Shopify.theme")

#  Shopify.theme = {"name":"Dawn","id":1234,"theme_store_id":887,...}
#  → capture the current theme NAME (goes to the 'Current Theme' column).
SHOPIFY_THEME_RE = re.compile(
    r'Shopify\.theme\s*=\s*\{[^{}]{0,300}?"name"\s*:\s*"((?:\\.|[^"\\]){1,80})"',
)


def is_shopify_html(html: str) -> bool:
    return any(m in html for m in SHOPIFY_MARKERS)


def extract_theme_name(html: str) -> str:
    m = SHOPIFY_THEME_RE.search(html)
    if not m:
        return ""
    name = m.group(1)
    # unescape \" and \/ that appear inside JSON string literals
    name = name.replace('\\"', '"').replace("\\/", "/").replace("\\\\", "\\")
    return name.strip()[:60]

# ═══════════════════════════════════════════════════════════════════
#   🚫  FILTERING CONSTANTS
# ═══════════════════════════════════════════════════════════════════

FREE_EMAIL_DOMAINS = {
    "gmail.com","yahoo.com","hotmail.com","outlook.com","icloud.com",
    "mail.com","protonmail.com","aol.com","yandex.ru","yandex.com",
    "zoho.com","live.com","msn.com","me.com","mac.com","inbox.com",
    "gmx.com","mail.ru","bk.ru","list.ru","inbox.ru","rambler.ru",
}

# Obvious placeholder / template / tracking addresses — hard reject.
PLACEHOLDER_RE = re.compile(
    r"your-?company|yourcompany|your-?domain|example\.|domain\.com"
    r"|email\.com|sample\.|test@|@test\.|name@|firstname|lastname"
    r"|user@|someone@|sentry|wixpress|\.png@|\.jpg@|@2x|sentry-next"
    r"|@sentry\.|@w3\.|@schema\.|@wordpress\.|@shopify\.com|@apple\.com"
    r"|@google\.com|@amazon\.com|@microsoft\.com|@facebook\.com"
    r"|sarah@startup\.com|john@|jane@|hello@example",
    re.IGNORECASE,
)

BAD_EMAIL_RE = re.compile(
    r"noreply|no-reply|donotreply|do-not-reply|unsubscribe"
    r"|privacy@|legal@|abuse@|spam@|postmaster@"
    r"|hostmaster@|mailer-daemon@|webmaster@",
    re.IGNORECASE,
)

# ─── Fresh filters ported from influencer_scraper v4 ─────────────

# domain "TLD" that is actually a file extension → asset filename, not email
# (e.g. "logo-27px@2x.png" matches the email regex)
FILE_EXT_TLDS = {
    "png", "jpg", "jpeg", "gif", "webp", "svg", "ico", "avif", "heic",
    "css", "js", "json", "map", "html", "htm", "php", "asp", "aspx",
    "mp4", "mov", "webm", "mp3", "wav", "pdf", "zip", "gz",
    "woff", "woff2", "ttf", "eot", "otf", "tld",
}

# placeholder + service local parts — useless for merchant outreach
BLOCKED_LOCALS = {
    "noreply", "no-reply", "donotreply", "do-not-reply", "webmaster",
    "postmaster", "abuse", "unsubscribe", "mailer-daemon", "bounce",
    "bounces", "billing", "legal", "privacy", "security", "admin",
    "administrator", "test", "testing",
    # service inboxes — not decision-makers
    "support", "help", "helpdesk", "service", "services",
    "customerservice", "customercare", "careers", "jobs", "hr",
    "recruiting", "recruitment", "news", "newsletter", "press", "media",
    # placeholder locals seen in article bodies / templates
    "your", "you", "yourname", "youremail", "your.email", "name",
    "firstname", "lastname", "first.last", "john.doe", "jane.doe",
    "user", "username", "someone", "email", "mail", "example", "sample",
    "blog",
}

# Big corporate media / SaaS blogs (from influencer v4): pass activity
# gates easily but are never a small merchant — skip their URLs outright.
BIG_PUBLISHER_DOMAINS = frozenset({
    "techradar.com", "itpro.com", "tomsguide.com", "pcmag.com", "cnet.com",
    "zdnet.com", "theverge.com", "wired.com", "engadget.com", "mashable.com",
    "digitaltrends.com", "lifehacker.com", "forbes.com", "businessinsider.com",
    "techcrunch.com", "entrepreneur.com", "inc.com", "fastcompany.com",
    "ycombinator.com", "industrydive.com", "marketingdive.com", "retaildive.com",
    "hubspot.com", "salesforce.com", "semrush.com", "ahrefs.com", "moz.com",
    "bigcommerce.com", "wix.com", "squarespace.com", "hostinger.com",
    "bluehost.com", "kinsta.com", "cloudways.com", "shopify.dev",
    "teamblind.com", "hackerone.com", "capterra.com", "g2.com",
    "trustpilot.com", "statista.com", "investopedia.com", "nerdwallet.com",
})

# Enterprise brands & marketplaces — pitching UTD themes to them is
# pointless (own dev teams / not on Shopify / marketplaces). Hard skip.
BIG_BRAND_DOMAINS = frozenset({
    # marketplaces & retail giants
    "amazon.com", "etsy.com", "ebay.com", "walmart.com", "target.com",
    "aliexpress.com", "alibaba.com", "temu.com", "wish.com", "shein.com",
    "wayfair.com", "costco.com", "homedepot.com", "lowes.com",
    "bestbuy.com", "asos.com", "zalando.com", "farfetch.com",
    "net-a-porter.com", "ssense.com", "nordstrom.com", "macys.com",
    # apparel / footwear giants
    "nike.com", "adidas.com", "puma.com", "underarmour.com",
    "newbalance.com", "zara.com", "hm.com", "uniqlo.com", "gap.com",
    "oldnavy.com", "levi.com", "lululemon.com",
    # luxury houses
    "gucci.com", "louisvuitton.com", "chanel.com", "dior.com",
    "prada.com", "burberry.com",
    # tech / beauty giants & huge DTC (own teams, custom themes)
    "apple.com", "samsung.com", "sony.com", "dyson.com",
    "sephora.com", "ulta.com", "ikea.com",
    "gymshark.com", "allbirds.com", "fashionnova.com",
    "kyliecosmetics.com", "colourpop.com",
})

# Directories, socials, review hubs — never a store homepage.
DIRECTORY_DOMAINS = {
    "clutch.co", "goodfirms.co", "designrush.com", "sortlist.com",
    "expertise.com", "upcity.com", "trustpilot.com", "yelp.com",
    "linkedin.com", "twitter.com", "x.com", "facebook.com", "instagram.com",
    "youtube.com", "tiktok.com", "pinterest.com", "reddit.com", "quora.com",
    "medium.com", "substack.com", "github.com", "wikipedia.org",
    "shopify.com", "themes.shopify.com", "apps.shopify.com",
    "upwork.com", "fiverr.com", "crunchbase.com", "producthunt.com",
}

GOV_EDU_TLDS = {
    ".gov",".mil",".edu",".ac.uk",".gov.uk",".gov.au",
    ".gov.ca",".gc.ca",".gouv.fr",".gob.es",".gov.ie",
}

# Page-path fragments that mean "article/listicle, not a store homepage".
# NB: /collections/ and /products/ are NOT junk here — they are exactly
# the Shopify URL patterns our queries hunt for (we normalise to homepage).
JUNK_PATH = [
    "/blog/", "/blogs/", "/news/", "/article", "/sitemap", "/html-sitemap",
    "/tag/", "/category/", "/author/", "/wp-json", "/feed",
]
JUNK_TITLE = [
    "sitemap", "top 10", "top 20", "best ", " vs ", "how to",
    "ultimate guide", "listicle", "privacy policy", "terms of",
    "404", "not found", "page not found", "blog -", "- blog",
]

# Signals that a domain is dead / parked / for-sale / suspended.
DEAD_SIGNALS = [
    "domain is for sale", "buy this domain", "this domain is for sale",
    "domain for sale", "is for sale", "parked free", "sedoparking",
    "this domain is parked", "domain parking", "hugedomains",
    "account suspended", "this account has been suspended",
    "site temporarily unavailable", "website is temporarily unavailable",
    "default web page", "welcome to nginx", "apache2 ubuntu default",
    "future home of something", "index of /",
    # Shopify-specific dead ends
    "this store is unavailable", "sorry, this shop is currently unavailable",
    "opening soon", "be right back",
]

# ─── Email priority (role quality, merchant edition) ─────────────
EMAIL_PRIORITY = [
    "contact", "hello", "hi", "hey", "info", "team", "shop", "store",
    "orders", "office", "connect", "bonjour", "care", "wholesale",
]
EMAIL_DEPRIO = {
    "sales", "billing", "invoice", "jobs", "career", "careers", "hr",
    "recruiting", "newsletter", "subscribe", "admin", "accounts",
    "accounting", "finance", "noreply", "returns",
}

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]{1,64}@[a-zA-Z0-9.\-]{1,253}\.[a-zA-Z]{2,12}",
    re.IGNORECASE,
)

# Generic multi-part TLDs needed to compute the registrable domain.
# "myshopify.com" is treated like a multi-part TLD so each store's
# foo.myshopify.com counts as its own domain (not one shared domain).
MULTI_TLDS = {
    "co.uk","org.uk","gov.uk","ac.uk","com.au","net.au","org.au",
    "co.nz","com.br","co.za","co.in","com.sg","com.hk","co.jp",
    "com.mx","com.tr","co.il","myshopify.com",
}

# ═══════════════════════════════════════════════════════════════════
#   📋  LOGGING
# ═══════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════
#   🌐  DOMAIN HELPERS
# ═══════════════════════════════════════════════════════════════════

def registrable_domain(host_or_url: str) -> str:
    """example.com from www.sub.example.com; handles co.uk-style TLDs
    and keeps foo.myshopify.com distinct per store."""
    h = host_or_url.strip().lower()
    if "://" in h:
        h = urlparse(h).netloc
    h = h.split("/")[0].split("@")[-1].replace("www.", "")
    if not h or "." not in h:
        return h
    parts = h.split(".")
    last2 = ".".join(parts[-2:])
    last3 = ".".join(parts[-3:]) if len(parts) >= 3 else ""
    if last2 in MULTI_TLDS and len(parts) >= 3:
        return ".".join(parts[-3:])
    return last2


def root_url(url: str) -> str:
    """Normalise any deep URL (e.g. /collections/all) to its homepage."""
    try:
        p = urlparse(url)
        if not p.scheme or not p.netloc:
            return ""
        return f"{p.scheme}://{p.netloc}"
    except Exception:
        return ""

# ═══════════════════════════════════════════════════════════════════
#   📨  MX RESOLVER  (DNS-over-HTTPS — works without port 25)
# ═══════════════════════════════════════════════════════════════════

class MXResolver:
    """
    Confirms a domain can receive email by checking for MX (or A as a
    fallback) records via DNS-over-HTTPS. Uses Cloudflare then Google,
    so it works even on networks that block outbound SMTP (port 25).
    Results are cached per domain for the whole run.
    """

    _DOH = [
        ("https://cloudflare-dns.com/dns-query", {"accept": "application/dns-json"}),
        ("https://dns.google/resolve", {}),
    ]

    def __init__(self):
        self.cache = {}     # domain -> bool
        self.http  = requests.Session()

    def _query(self, domain: str, rtype: str) -> list:
        for base, hdrs in self._DOH:
            try:
                r = self.http.get(
                    base, params={"name": domain, "type": rtype},
                    headers=hdrs, timeout=10,
                )
                if r.status_code == 200:
                    return r.json().get("Answer", []) or []
            except Exception:
                continue
        return None  # network failure (distinct from "no records")

    def has_mail(self, domain: str) -> bool:
        domain = registrable_domain(domain)
        if not domain:
            return False
        if domain in self.cache:
            return self.cache[domain]

        mx = self._query(domain, "MX")
        if mx is None:                       # DoH failed → don't punish domain
            return True
        ok = len(mx) > 0
        if not ok:                           # no MX: many small sites use A for mail
            a = self._query(domain, "A")
            ok = bool(a)
        self.cache[domain] = ok
        return ok

# ═══════════════════════════════════════════════════════════════════
#   📧  EMAIL UTILITIES
# ═══════════════════════════════════════════════════════════════════

def is_valid_email(em: str) -> bool:
    em = em.lower().strip()
    if not em or "@" not in em or len(em) > 254:
        return False
    local, domain = em.rsplit("@", 1)
    if not local or not domain or "." not in domain:
        return False
    # malformed edges the broad regex lets through (CSS/JS artifacts)
    if local[0] in ".-_" or local[-1] in ".-_":
        return False
    if any(lbl.startswith("-") or lbl.endswith("-") or not lbl
           for lbl in domain.split(".")):
        return False
    if domain in FREE_EMAIL_DOMAINS:
        return False
    # asset filenames: "TLD" is really a file extension ("logo@2x.png")
    if domain.rsplit(".", 1)[-1] in FILE_EXT_TLDS:
        return False
    # retina/asset local parts: "...-27px", "300x250", pure hex hashes
    if re.search(r"\d+px$|\d+x\d+$", local) or re.fullmatch(r"[0-9a-f]{8,}", local):
        return False
    if local in BLOCKED_LOCALS:
        return False
    if re.match(r"^u[0-9a-f]{3,5}$", local):   # JS escape artifact
        return False
    if PLACEHOLDER_RE.search(em) or BAD_EMAIL_RE.search(em):
        return False
    tld_len = len(domain.split(".")[-1])
    return 2 <= tld_len <= 12


def extract_emails(text: str) -> set:
    return {e.lower() for e in EMAIL_RE.findall(text) if is_valid_email(e)}


def email_role_score(local: str) -> int:
    base = local.split("@")[0].replace(".", "").replace("-", "").replace("_", "")
    first = local.split("@")[0].split(".")[0]
    if base in EMAIL_PRIORITY:
        return 25
    if first in EMAIL_DEPRIO:
        return 6
    return 16  # personal name e.g. anna@ — fine for outreach


def pick_best_email(emails: set, site_domain: str) -> str:
    """Prefer on-domain, role-priority addresses."""
    on_domain = [e for e in emails
                 if registrable_domain(e.split("@")[-1]) == site_domain]
    pool = on_domain or list(emails)
    if not pool:
        return ""
    return sorted(pool, key=lambda e: (-email_role_score(e.split("@")[0]), e))[0]

# ═══════════════════════════════════════════════════════════════════
#   🏪  URL PRE-FILTER
# ═══════════════════════════════════════════════════════════════════

def is_store_url(url: str) -> bool:
    """Reject directories, enterprise brands, publishers, gov/edu, articles."""
    try:
        p = urlparse(url)
        domain = p.netloc.lower().replace("www.", "")
        path   = p.path.lower()
    except Exception:
        return False
    if not domain:
        return False
    for pool in (BIG_BRAND_DOMAINS, BIG_PUBLISHER_DOMAINS, DIRECTORY_DOMAINS):
        if any(domain == b or domain.endswith("." + b) for b in pool):
            return False
    if any(domain.endswith(t) for t in GOV_EDU_TLDS):
        return False
    if any(j in path for j in JUNK_PATH):
        return False
    return True

# ═══════════════════════════════════════════════════════════════════
#   🏷️  INDUSTRY CLASSIFIER  (keyword scoring, PROFILE-driven)
# ═══════════════════════════════════════════════════════════════════

def classify_industry(title: str, meta_desc: str, page_text: str) -> str:
    """Pick the industry with the most marker hits in title+meta+text.
    Title/meta hits weigh 3× (they name the store's own business)."""
    head = f"{title} {meta_desc}".lower()
    body = (page_text or "")[:4000].lower()
    best_label, best_score = PROFILE["industry_default"], 0
    for label, markers in PROFILE["industries"].items():
        score = 0
        for kw in markers:
            if kw in head:
                score += 3
            elif kw in body:
                score += 1
        if score > best_score:
            best_label, best_score = label, score
    return best_label if best_score >= 2 else PROFILE["industry_default"]

# ═══════════════════════════════════════════════════════════════════
#   ✅  VALIDATORS  (is it a real STORE, and which industry?)
# ═══════════════════════════════════════════════════════════════════

class RuleValidator:
    def classify(self, store: str, url: str, page_text: str) -> str:
        blob = (store + " " + url + " " + page_text[:4000]).lower()
        neg  = sum(1 for kw in PROFILE["negative"] if kw in blob)
        pos  = sum(1 for kw in PROFILE["positive"] if kw in blob)
        if neg >= 2:              return "reject"
        if neg >= 1 and pos == 0: return "reject"
        if pos >= 3:              return "accept"
        if pos >= 1 and neg == 0: return "accept"
        return "uncertain"


class ClaudeValidator:
    """Optional Claude check: confirms it's a real store AND may refine
    the industry label (replies 'ACCEPT <Industry>' or 'REJECT')."""

    _API = "https://api.anthropic.com/v1/messages"
    _MDL = "claude-haiku-4-5-20251001"

    def __init__(self, key: str):
        self.key    = (key or "").strip()
        self._cache = {}

    def enabled(self) -> bool:
        return bool(self.key)

    def classify(self, store: str, url: str, snippet: str) -> tuple:
        """Returns (verdict, industry): verdict ∈ accept/reject/unknown."""
        if not self.key:
            return "unknown", ""
        dom = registrable_domain(url)
        if dom in self._cache:
            return self._cache[dom]
        try:
            r = self._call(store, url, snippet[:600])
            self._cache[dom] = r
            return r
        except Exception as e:
            log.debug(f"Claude error: {e}")
            return "unknown", ""

    def _call(self, store: str, url: str, snippet: str) -> tuple:
        hdrs = {"x-api-key": self.key, "anthropic-version": "2023-06-01",
                "content-type": "application/json"}
        labels = ", ".join(list(PROFILE["industries"].keys())
                           + [PROFILE["industry_default"]])
        prompt = (
            f"{PROFILE['claude_question']}\n\n"
            f"Store: {store}\nURL: {url}\nPage excerpt: {snippet}\n\n"
            "Reply with exactly one line:\n"
            f"ACCEPT <industry>  — yes, it is a real store; <industry> is one of: {labels}.\n"
            "REJECT — no (agency, app/theme vendor, marketplace, blog, "
            "service company, or unrelated)."
        )
        body = {"model": self._MDL, "max_tokens": 20,
                "messages": [{"role": "user", "content": prompt}]}
        resp = requests.post(self._API, headers=hdrs, json=body, timeout=22)
        if resp.status_code == 200:
            txt = resp.json()["content"][0]["text"].strip()
            up  = txt.upper()
            if "REJECT" in up:
                return "reject", ""
            if "ACCEPT" in up:
                rest = txt[up.index("ACCEPT") + 6:].strip(" :—-")
                industry = ""
                for lbl in list(PROFILE["industries"].keys()) + [PROFILE["industry_default"]]:
                    if lbl.lower() in rest.lower():
                        industry = lbl
                        break
                return "accept", industry
        elif resp.status_code == 529:
            time.sleep(8)
        return "unknown", ""


class StoreValidator:
    def __init__(self, anthropic_key: str = ""):
        self.rules  = RuleValidator()
        self.claude = ClaudeValidator(anthropic_key)
        self.stats  = {"accepted": 0, "rejected": 0,
                       "claude_calls": 0, "uncertain_dropped": 0}

    def validate(self, store: str, url: str, page_text: str,
                 industry: str) -> tuple:
        """Returns (ok, method, industry) — Claude may refine the industry."""
        result = self.rules.classify(store, url, page_text)
        if result == "reject":
            self.stats["rejected"] += 1
            return False, "rules", industry
        if result == "accept":
            self.stats["accepted"] += 1
            return True, "rules", industry
        if self.claude.enabled():
            self.stats["claude_calls"] += 1
            cr, c_ind = self.claude.classify(store, url, page_text)
            if cr == "accept":
                self.stats["accepted"] += 1
                return True, "claude", (c_ind or industry)
            if cr == "reject":
                self.stats["rejected"] += 1
                return False, "claude", industry
        self.stats["uncertain_dropped"] += 1
        return False, "uncertain", industry

# ═══════════════════════════════════════════════════════════════════
#   🔎  CONTACT VERIFIER  (is the store alive & email deliverable?)
# ═══════════════════════════════════════════════════════════════════

class ContactVerifier:
    """
    Produces (status, score, reason) for an email. Free signals only —
    no port-25 SMTP needed. Identical gauntlet to b2b v4.
    """

    def __init__(self, mx: MXResolver):
        self.mx = mx
        self.stats = {"no_mx": 0, "mismatch": 0, "bad_format": 0,
                      "dead_site": 0, "live": 0, "risky": 0}

    def verify(self, email: str, site_domain: str,
               page_text: str, site_alive: bool) -> tuple:
        email = email.lower().strip()

        # 1. Format / placeholder
        if not is_valid_email(email):
            self.stats["bad_format"] += 1
            return "dead", 0, "bad_format"

        edom = registrable_domain(email.split("@")[-1])

        # 2. Email must live on the store's OWN domain (kills 3rd-party junk).
        #    Exception: a store still on foo.myshopify.com can't host mail
        #    there — accept its corporate-domain email instead.
        if site_domain and edom != site_domain:
            if not site_domain.endswith("myshopify.com"):
                self.stats["mismatch"] += 1
                return "dead", 0, "domain_mismatch"

        # 3. Domain must be able to receive mail (MX via DoH)
        if not self.mx.has_mail(edom):
            self.stats["no_mx"] += 1
            return "dead", 0, "no_mx"

        # 4. Website must be alive (checked upstream during scrape)
        if not site_alive:
            self.stats["dead_site"] += 1
            return "dead", 0, "dead_site"

        # ── Scoring (0-100) ──────────────────────────────────────
        score = 40                                   # MX present + on-domain
        score += 20                                  # site live
        score += email_role_score(email.split("@")[0])   # role quality 6-25

        # freshness bonus: recent year somewhere on the page
        try:
            yrs = [int(y) for y in re.findall(r"\b(20[12]\d)\b", page_text or "")]
            this_year = datetime.now().year
            if yrs and max(yrs) >= this_year - 1:
                score += 15
        except Exception:
            pass

        score = min(score, 100)
        status = "live" if score >= MIN_SCORE else "risky"
        self.stats[status] += 1
        return status, score, "ok"

# ═══════════════════════════════════════════════════════════════════
#   ✉️  MAILBOX VERIFIER  (optional free-tier API — real RCPT check)
# ═══════════════════════════════════════════════════════════════════

class MailboxVerifier:
    """
    Confirms a specific mailbox exists via a hosted verification API
    (MillionVerifier or Reoon). Multiple API keys rotate round-robin;
    an exhausted/invalid key drops out for the rest of the run.
    Statuses: 'deliverable' | 'catch_all' | 'undeliverable' | 'unknown' | 'off'
    """

    def __init__(self, provider: str, keys):
        self.provider = (provider or "").strip().lower()
        if isinstance(keys, str):
            keys = [keys]
        self.keys = [k.strip() for k in (keys or []) if k and k.strip()]
        self.idx  = 0
        self.dead = set()              # indices of exhausted/invalid keys
        self.http = requests.Session()
        self.cache = {}
        self.per_key = [0] * len(self.keys)
        self.stats = {"deliverable": 0, "catch_all": 0,
                      "undeliverable": 0, "unknown": 0, "calls": 0}

    def enabled(self) -> bool:
        return bool(self.provider and self.keys)

    def exhausted(self) -> bool:
        return self.enabled() and len(self.dead) >= len(self.keys)

    def _next_key_idx(self):
        for _ in range(len(self.keys)):
            i = self.idx % len(self.keys)
            self.idx += 1
            if i not in self.dead:
                return i
        return None

    def verify(self, email: str) -> str:
        if not self.enabled():
            return "off"
        email = email.lower().strip()
        if email in self.cache:
            return self.cache[email]

        res = "unknown"
        for _ in range(len(self.keys)):
            i = self._next_key_idx()
            if i is None:
                log.warning("  ✉️  all mailbox keys exhausted — skipping check")
                break
            key = self.keys[i]
            self.per_key[i] += 1
            try:
                if self.provider == "reoon":
                    status, failed, exhaust = self._reoon(email, key)
                elif self.provider == "millionverifier":
                    status, failed, exhaust = self._millionverifier(email, key)
                else:
                    log.warning(f"  unknown mailbox provider '{self.provider}'")
                    break
            except Exception as e:
                log.debug(f"  mailbox API error: {e}")
                status, failed, exhaust = "unknown", True, False

            if exhaust:
                self.dead.add(i)
                log.warning(f"  🔄 mailbox key #{i+1} (…{key[-4:]}) "
                            f"exhausted/invalid — dropped from rotation")
            if failed:
                continue          # try next key
            res = status
            break

        self.stats["calls"] += 1
        self.stats[res] = self.stats.get(res, 0) + 1
        self.cache[email] = res
        return res

    # ── MillionVerifier: api.millionverifier.com/api/v3 ─────────
    def _millionverifier(self, email: str, key: str):
        r = self.http.get(
            "https://api.millionverifier.com/api/v3/",
            params={"api": key, "email": email, "timeout": 20},
            timeout=35,
        )
        if r.status_code in (401, 403):
            return "unknown", True, True
        if r.status_code == 429:
            return "unknown", True, False     # rate-limited, key still good
        if r.status_code != 200:
            return "unknown", True, False
        d = r.json()
        err = str(d.get("error", "")).lower()
        if err:
            kill = any(w in err for w in ("credit", "api key", "invalid", "key"))
            return "unknown", True, kill
        res = str(d.get("result", "")).lower()
        credits = d.get("credits")
        out_of_credits = isinstance(credits, (int, float)) and credits <= 0
        if res == "ok":
            return "deliverable", False, out_of_credits
        if res in ("catch_all", "catch-all"):
            return "catch_all", False, out_of_credits
        if res in ("invalid", "disposable", "error"):
            return "undeliverable", False, out_of_credits
        return "unknown", False, out_of_credits

    # ── Reoon: emailverifier.reoon.com/api/v1 ───────────────────
    def _reoon(self, email: str, key: str):
        r = self.http.get(
            "https://emailverifier.reoon.com/api/v1/verify",
            params={"email": email, "key": key, "mode": "power"},
            timeout=35,
        )
        if r.status_code in (401, 403):
            return "unknown", True, True
        if r.status_code != 200:
            return "unknown", True, False
        d = r.json()
        st = str(d.get("status", "")).lower()
        if d.get("is_safe_to_send") is True or st in ("safe", "valid"):
            return "deliverable", False, False
        if st in ("invalid", "disabled", "disposable", "spamtrap", "undeliverable"):
            return "undeliverable", False, False
        if st in ("catch_all", "catch-all", "accept_all"):
            return "catch_all", False, False
        return "unknown", False, False

# ═══════════════════════════════════════════════════════════════════
#   🔄  DYNAMIC QUERY GENERATOR  (<niche> × <intent> × <geo>)
# ═══════════════════════════════════════════════════════════════════

class QueryGenerator:
    """Unlimited unique, Shopify-store-targeted Google queries
    (deterministic — same counter always yields the same query)."""

    # Same geo pool as b2b (US/EU/UK/CA/AU cities; blanks = no-geo queries).
    _LOC = [
        "", "", "", "", "", "",
        # North America
        "USA", "New York", "Los Angeles", "Chicago", "San Francisco",
        "Austin", "Seattle", "Boston", "Miami", "Denver", "Dallas",
        "Atlanta", "Portland", "San Diego", "Phoenix", "Houston",
        "Philadelphia", "Minneapolis", "Nashville", "Charlotte",
        "Orlando", "Tampa", "Las Vegas", "Columbus", "Kansas City",
        "Indianapolis", "Detroit", "Pittsburgh", "Salt Lake City",
        "Canada", "Toronto", "Vancouver", "Montreal", "Calgary", "Ottawa",
        # UK & Ireland
        "UK", "London", "Manchester", "Birmingham", "Bristol",
        "Leeds", "Glasgow", "Edinburgh", "Liverpool", "Nottingham",
        "Ireland", "Dublin",
        # Europe
        "Germany", "Berlin", "Munich", "Hamburg", "Cologne",
        "Netherlands", "Amsterdam", "Rotterdam", "Utrecht",
        "France", "Paris", "Lyon", "Spain", "Madrid", "Barcelona",
        "Valencia", "Italy", "Milan", "Rome", "Portugal", "Lisbon",
        "Porto", "Poland", "Warsaw", "Krakow", "Wroclaw",
        "Czech Republic", "Prague", "Austria", "Vienna",
        "Switzerland", "Zurich", "Geneva", "Sweden", "Stockholm",
        "Denmark", "Copenhagen", "Norway", "Oslo", "Finland", "Helsinki",
        "Belgium", "Brussels", "Antwerp", "Romania", "Bucharest",
        "Bulgaria", "Sofia", "Croatia", "Zagreb", "Greece", "Athens",
        "Estonia", "Tallinn", "Latvia", "Riga", "Lithuania", "Vilnius",
        "Ukraine", "Kyiv", "Lviv", "Hungary", "Budapest", "Serbia", "Belgrade",
        # APAC & other
        "Australia", "Sydney", "Melbourne", "Brisbane", "Perth", "Adelaide",
        "Singapore", "New Zealand", "Auckland", "Dubai", "UAE",
        "South Africa", "Cape Town", "Johannesburg",
    ]

    # Cut marketplace/article noise from the SERP up front.
    _EXCLUDE = '-amazon -etsy -ebay -aliexpress -"top 10" -"best" -wikipedia'

    # Intent patterns tuned to surface actual Shopify storefronts:
    #  • "powered by shopify" — the classic footer credit
    #  • site:myshopify.com — stores still on the default subdomain
    #  • inurl:collections / inurl:products — canonical Shopify URL paths
    #  {n} = niche, {loc} = optional geo, {ex} = exclusions.
    _INTENTS = [
        '"powered by shopify" {n} {loc}{ex}',
        '"powered by shopify" {n} store {ex}',
        'site:myshopify.com {n}',
        'site:myshopify.com {n} shop',
        '{n} "online store" inurl:collections {loc}{ex}',
        '{n} shop inurl:products {loc}{ex}',
        '{n} store inurl:collections {ex}',
        '{n} brand "shop now" inurl:products {ex}',
        '{n} boutique "add to cart" {loc}{ex}',
        '{n} "free shipping" inurl:collections {ex}',
    ]

    def __init__(self, counter: int, used_hashes: set):
        self.counter     = counter
        self.used_hashes = used_hashes

    def _hash(self, q: str) -> str:
        return hashlib.md5(q.lower().strip().encode()).hexdigest()[:14]

    def _build(self) -> str:
        rng = random.Random(self.counter)
        self.counter += 1
        niche  = rng.choice(PROFILE["niches"])
        intent = rng.choice(self._INTENTS)
        loc    = rng.choice(self._LOC)
        loc_s  = f"{loc} " if loc else ""
        q = intent.format(n=niche, loc=loc_s, ex=self._EXCLUDE)
        return re.sub(r"\s{2,}", " ", q).strip()

    def next_batch(self, n: int) -> list:
        queries, attempts = [], 0
        while len(queries) < n and attempts < n * 300:
            q = self._build()
            h = self._hash(q)
            if h not in self.used_hashes and len(q) > 8:
                self.used_hashes.add(h)
                queries.append(q)
            attempts += 1
        return queries

# ═══════════════════════════════════════════════════════════════════
#   🔑  GOOGLE SEARCH CLIENT  (rotating keys)
# ═══════════════════════════════════════════════════════════════════

class GoogleClient:
    _BASE = "https://www.googleapis.com/customsearch/v1"

    def __init__(self, pairs: list):
        self.pairs    = pairs
        self.idx      = 0
        self.failures = [0] * len(pairs)
        self.dead     = set()          # pairs out of daily quota (429/403)
        self.calls    = 0

    def exhausted(self) -> bool:
        """True once every key pair is out of quota for the day. The harvester
        checks this and stops early instead of storming thousands of 429s."""
        return len(self.dead) >= len(self.pairs)

    def _advance(self):
        for _ in range(len(self.pairs)):
            self.idx = (self.idx + 1) % len(self.pairs)
            if self.idx not in self.dead:
                return

    def _drop(self, reason: str):
        self.dead.add(self.idx)
        log.warning(f"  🔻 API pair #{self.idx} dropped ({reason}); "
                    f"{len(self.pairs) - len(self.dead)} live")
        self._advance()

    def _rotate(self, reason: str):
        self.failures[self.idx] += 1
        self._advance()
        time.sleep(API_ROT_WAIT)

    @staticmethod
    def _is_transient_rate(resp) -> bool:
        """True ONLY for a transient per-minute/second rate limit — pause and try
        another key, keeping this one alive. Everything else (daily quota,
        accessNotConfigured, no-access, invalid, disabled) is permanent for the
        run → drop the key so we don't grind on dead keys."""
        try:
            body = resp.text.lower()
        except Exception:
            return False
        if any(k in body for k in ("per day", "daily", "dailylimitexceeded")):
            return False
        return any(k in body for k in (
            "ratelimitexceeded", "userratelimitexceeded", "user rate limit",
            "per minute", "perminute", "too many requests"))

    def search(self, query: str, start: int = 1) -> list:
        if self.exhausted():
            return []
        for _ in range(len(self.pairs)):
            if self.exhausted():
                break
            if self.idx in self.dead:
                self._advance(); continue
            pair = self.pairs[self.idx]
            try:
                resp = requests.get(
                    self._BASE,
                    params={"key": pair["api_key"], "cx": pair["cse_id"],
                            "q": query, "num": RESULTS_PER_QUERY,
                            "start": start,
                            "hl": "en", "lr": "lang_en"},
                    timeout=15,
                )
                self.calls += 1
                if resp.status_code == 200:
                    return [it["link"] for it in resp.json().get("items", []) if "link" in it]
                if resp.status_code in (429, 403):
                    if self._is_transient_rate(resp):
                        # Momentary rate limit — pause and try another key, but
                        # keep this one alive (a burst must not burn good keys).
                        self._rotate(f"HTTP {resp.status_code} rate")
                    else:
                        # Daily quota OR permanent block — drop for the run.
                        self._drop(f"HTTP {resp.status_code}")
                elif resp.status_code == 400:
                    return []
                else:
                    self._rotate(f"HTTP {resp.status_code}")
            except requests.Timeout:
                self._rotate("timeout")
            except requests.RequestException as e:
                self._rotate(str(e)[:40])
        if self.exhausted():
            log.error("  ❌ All API pairs out of quota — stopping search.")
        return []

# ═══════════════════════════════════════════════════════════════════
#   🛍️  STORE WEBSITE SCRAPER  (homepage-anchored, Shopify-gated)
# ═══════════════════════════════════════════════════════════════════

def _headers() -> dict:
    return {
        "User-Agent":      random.choice(USER_AGENTS),
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate",
        "Connection":      "keep-alive",
    }

# Shopify stores keep contact pages under /pages/… by convention.
_SUBPAGES = ["/pages/contact", "/pages/contact-us", "/pages/about",
             "/pages/about-us", "/contact", "/policies/contact-information"]


def _store_name(soup: BeautifulSoup, url: str) -> str:
    og = soup.find("meta", property="og:site_name")
    if og and og.get("content", "").strip():
        return og["content"].strip()[:80]
    t = soup.find("title")
    if t:
        txt = t.get_text(strip=True)
        for sep in [" | ", " – ", " — ", " - ", " · ", " :: "]:
            if sep in txt:
                return txt.split(sep)[0].strip()[:80]
        return txt[:80]
    return registrable_domain(url)


def _meta_description(soup: BeautifulSoup) -> str:
    m = soup.find("meta", attrs={"name": "description"})
    if m and m.get("content"):
        return m["content"].strip()[:400]
    m = soup.find("meta", property="og:description")
    if m and m.get("content"):
        return m["content"].strip()[:400]
    return ""


def _looks_like_junk(name: str) -> bool:
    n = name.lower()
    return any(k in n for k in JUNK_TITLE)


def scrape_store(home_url: str, session: requests.Session,
                 visited: set) -> dict:
    """
    Scrape a store HOMEPAGE (+ a few contact subpages) for one best email.
    Returns dict: {email, store, domain, industry, theme, page_text,
                   alive, shopify, url}.
    'alive'   is False if the site is down/parked/suspended.
    'shopify' is False if no Shopify marker was found (HARD GATE upstream).
    """
    out = {"email": "", "store": "", "domain": "", "industry": "",
           "theme": "", "page_text": "", "alive": False, "shopify": False,
           "url": home_url}

    root = root_url(home_url)
    if not root:
        return out
    out["domain"] = registrable_domain(root)

    mailto_emails, text_emails = set(), set()
    tried = 0
    pages = [root]                      # subpages appended after redirect

    while pages:
        url = pages.pop(0)
        if not url:
            continue
        uh = hash_key(url)                # visited stores URL hashes (repo is public)
        if uh in visited:
            continue
        if tried >= MAX_SUBPAGES + 1:
            break
        visited.add(uh)
        tried += 1
        try:
            resp = session.get(url, headers=_headers(), timeout=13,
                               allow_redirects=True)
            if resp.status_code != 200:
                continue
            if "text/html" not in resp.headers.get("Content-Type", ""):
                continue

            html = resp.text
            soup = BeautifulSoup(html, "html.parser")
            body = soup.get_text(" ", strip=True)

            if tried == 1:  # homepage: redirect, liveness, Shopify gate, identity
                # follow the final URL (myshopify.com often 301s to the
                # custom domain — that's the store's real address)
                final_root = root_url(resp.url) or root
                out["url"]    = final_root
                out["domain"] = registrable_domain(final_root)

                bl = body.lower()
                if len(bl) < 200 or any(s in bl for s in DEAD_SIGNALS):
                    log.info(f"    ⛔ dead/parked: {out['domain']}")
                    return out                       # alive stays False

                # ── HARD SHOPIFY GATE ────────────────────────────
                if not is_shopify_html(html):
                    log.info(f"    ⛔ not Shopify: {out['domain']}")
                    return out                       # shopify stays False
                out["shopify"]  = True
                out["alive"]    = True
                out["theme"]    = extract_theme_name(html)
                out["store"]    = _store_name(soup, final_root)
                out["page_text"]= body[:3000]
                out["industry"] = classify_industry(
                    soup.title.get_text(strip=True) if soup.title else "",
                    _meta_description(soup), body)

                pages = [final_root + sp for sp in _SUBPAGES]

            # theme may appear on subpages when homepage was heavily cached
            if not out["theme"]:
                out["theme"] = extract_theme_name(html)

            for a in soup.find_all("a", href=True):
                h = a["href"]
                if h.lower().startswith("mailto:"):
                    em = h[7:].split("?")[0].strip().lower()
                    if is_valid_email(em):
                        mailto_emails.add(em)
            text_emails.update(extract_emails(html))

            best = pick_best_email(mailto_emails, out["domain"])
            if best and email_role_score(best.split("@")[0]) == 25:
                break
            if tried > 1:
                time.sleep(DELAY_PAGE)
        except requests.Timeout:
            log.debug(f"  timeout: {url}")
        except Exception as e:
            log.debug(f"  fetch error {url}: {e}")

    out["email"] = (pick_best_email(mailto_emails, out["domain"])
                    or pick_best_email(text_emails, out["domain"]))
    return out

# ═══════════════════════════════════════════════════════════════════
#   📊  EXCEL OUTPUT  (local mirror of the sheet's A–H block)
# ═══════════════════════════════════════════════════════════════════

_DARK, _EVEN, _ODD = "0A0A0A", "F4F6EE", "FFFFFF"
COLS   = ["Email", "Store Name", "Website", "Industry",
          "Current Theme", "Platform", "Validated By", "Date Found"]
WIDTHS = [36, 30, 42, 24, 22, 12, 22, 18]


def _border():
    s = Side(style="thin", color="C5D5E8")
    return Border(left=s, right=s, top=s, bottom=s)


def init_excel(path: str):
    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path); wb.close(); return
        except Exception:
            os.remove(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shopify Store Contacts"
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color=_DARK, end_color=_DARK, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 26
    wb.save(path)
    log.info(f"📗 Excel created: {path}")


def _row_values(r: dict) -> list:
    return [r["email"], r["store"], r["url"], r["industry"],
            r["theme"], "Shopify", r["validated_by"], r["date"]]


def append_to_excel(path: str, records: list):
    if not records:
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    start = ws.max_row + 1
    for i, r in enumerate(records):
        row = start + i
        fc = _EVEN if row % 2 == 0 else _ODD
        fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
        for col, val in enumerate(_row_values(r), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(vertical="center")
            c.border = _border()
        ws.row_dimensions[row].height = 18
    wb.save(path)
    log.info(f"  📗 +{len(records)} rows  (file: {path})")

# ═══════════════════════════════════════════════════════════════════
#   🟢  GOOGLE SHEETS UPLOADER  (n8n-safe: writes ONLY columns A–H)
# ═══════════════════════════════════════════════════════════════════
#
#  Sheet 'Contacts' layout (the parser fills A–H; I–O belong to n8n):
#   A Email | B Store Name | C Website | D Industry | E Current Theme
#   F Platform | G Validated By | H Date Found | I Status | J Date Sent
#   K Thread ID | L Date Replied | M Suggested Themes | N Last Msg ID
#   O Notes

class SheetsUploader:
    def __init__(self, spreadsheet_id, worksheet_name, creds_file):
        self.spreadsheet_id = spreadsheet_id
        self.worksheet_name = worksheet_name
        self.creds_file     = creds_file
        self._ws = None; self._ok = False; self._sent = 0; self._next_row = None

    def connect(self) -> bool:
        if not self.spreadsheet_id:
            log.warning("⚠️  ECOM_SHEET_ID is not set; Excel only.")
            return False
        if not os.path.exists(self.creds_file):
            log.warning(f"⚠️  Sheets creds not found ({self.creds_file}); Excel only.")
            return False
        try:
            import gspread
        except ImportError:
            log.error("⚠️  gspread not installed; Excel only. "
                      "pip install gspread google-auth")
            return False
        try:
            gc = gspread.service_account(filename=self.creds_file)
            ss = gc.open_by_key(self.spreadsheet_id)
            try:
                self._ws = ss.worksheet(self.worksheet_name)
            except Exception:
                self._ws = ss.get_worksheet(0)
                log.warning(f"  tab '{self.worksheet_name}' not found — "
                            f"using first tab '{self._ws.title}'")
            if self._ws is None:
                log.error("  No worksheet found."); return False
            self._ok = True
            # Cache the append cursor ONCE. Re-reading the whole sheet on every
            # flush (get_all_values on 1000+ rows) burns the Sheets read quota
            # and was the cause of the 429s that silently dropped 218 rows.
            try:
                self._next_row = len(self._ws.get_all_values()) + 1
            except Exception:
                self._next_row = None
            log.info(f"  ✅ Google Sheets ready: tab='{self._ws.title}' in '{ss.title}'")
            return True
        except Exception as e:
            log.error(f"⚠️  Sheets connect error: {e}; Excel only.")
            return False

    def existing_emails(self) -> set:
        """Emails already in column A — folded into dedup at startup."""
        if not self._ok:
            return set()
        try:
            col = self._ws.col_values(1)
            return {str(v).strip().lower() for v in col[1:] if v and "@" in str(v)}
        except Exception as e:
            log.warning(f"  Sheets: could not read existing emails ({e})")
            return set()

    def append_rows(self, records: list) -> bool:
        """
        Writes ONLY columns A–H at the TRUE last row + 1 via update()
        (not append_rows — a blank row mid-sheet makes append_rows insert
        at the gap and shift the n8n columns; update() is gap-proof).
        Columns I–O (Status / Date Sent / Thread ID / …) are never touched.
        """
        if not self._ok or not self._ws or not records:
            return False
        rows = [_row_values(r) for r in records]
        for attempt in range(1, 6):
            try:
                # Use the cached cursor; only re-read the sheet if we don't have
                # one yet or a prior attempt in this call may have drifted it.
                if self._next_row is None:
                    self._next_row = len(self._ws.get_all_values()) + 1
                next_row = self._next_row
                # ws.update() never grows the grid — writing past row_count
                # fails with "exceeds grid limits" (this silently dropped 218
                # rows on 2026-07-12). Add rows first if the batch won't fit.
                need = next_row + len(rows) - 1
                if self._ws.row_count < need:
                    self._ws.add_rows(need - self._ws.row_count + 50)
                self._ws.update(
                    range_name=f"A{next_row}",
                    values=rows,
                    value_input_option="USER_ENTERED",
                )
                self._next_row = next_row + len(rows)
                self._sent += len(rows)
                log.info(f"  🟢 Sheets: +{len(rows)} rows at A{next_row} "
                         f"(session {self._sent})")
                return True
            except Exception as e:
                err = str(e).lower()
                if any(k in err for k in ("429", "quota", "rate")):
                    log.warning(f"  Sheets rate-limited (attempt {attempt}/5): {str(e)[:120]}")
                    self._next_row = None          # force a fresh cursor next try
                    time.sleep(20 * attempt)
                elif "401" in err or "403" in err:
                    log.error(f"  Sheets auth error: {e}"); self._ok = False; return False
                else:
                    log.warning(f"  Sheets write error (attempt {attempt}/5): {str(e)[:120]}")
                    self._next_row = None
                    time.sleep(6 * attempt)
        log.error(f"  Sheets: 5 attempts failed — batch of {len(rows)} NOT written.")
        return False

# ═══════════════════════════════════════════════════════════════════
#   💾  STATE
# ═══════════════════════════════════════════════════════════════════

def hash_key(s: str) -> str:
    """SHA256 of a normalised string. Used to store emails/domains in the
    committed data/*.json as one-way hashes (repo is PUBLIC) while keeping
    dedup working: hash(email) collides iff the emails are equal.
    Real emails/domains are still written to the private Google Sheet + Excel."""
    return hashlib.sha256(str(s).lower().strip().encode()).hexdigest()


def _default_state() -> dict:
    return {"version": "1.0", "query_counter": 0, "used_query_hashes": [],
            "emails": {}, "seen_domains": [], "session_count": 0,
            "created_at": datetime.now().isoformat(),
            "last_updated": datetime.now().isoformat()}


def load_state() -> dict:
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                s = json.load(f)
            for k, v in _default_state().items():
                s.setdefault(k, v)
            log.info(f"📂 State: {len(s['emails'])} contacts, "
                     f"query #{s['query_counter']}, {len(s['seen_domains'])} domains")
            return s
        except Exception as e:
            log.warning(f"State load error ({e}). Fresh start.")
    return _default_state()


def save_state(s: dict):
    s["last_updated"] = datetime.now().isoformat()
    if len(s["used_query_hashes"]) > 60000:
        s["used_query_hashes"] = s["used_query_hashes"][-50000:]
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(s, f, indent=2, ensure_ascii=False)


def load_visited() -> set:
    if os.path.exists(VISITED_FILE):
        try:
            with open(VISITED_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def save_visited(v: set):
    with open(VISITED_FILE, "w", encoding="utf-8") as f:
        json.dump(list(v)[-120000:], f, ensure_ascii=False)


def reset_state():
    for p in [STATE_FILE, VISITED_FILE]:
        if os.path.exists(p):
            os.remove(p)
    log.info("🗑️  State cleared.")

# ═══════════════════════════════════════════════════════════════════
#   🚀  HARVESTER
# ═══════════════════════════════════════════════════════════════════

class EcomHarvester:

    def __init__(self, anthropic_key: str = ""):
        self.state        = load_state()
        self.visited      = load_visited()
        self.seen_domains = set(self.state["seen_domains"])
        self.used_hashes  = set(self.state["used_query_hashes"])

        self.google    = GoogleClient(API_PAIRS)
        self.http      = requests.Session()
        self.http.max_redirects = 5
        self.mx        = MXResolver()
        self.classifier= StoreValidator(anthropic_key)
        self.verifier  = ContactVerifier(self.mx)
        self.mailbox   = MailboxVerifier(MAILBOX_PROVIDER, MAILBOX_API_KEYS)
        self.qgen      = QueryGenerator(self.state["query_counter"], self.used_hashes)

        self.running   = True
        self._buf      = []
        self._warned_nocredits = False
        self.stats     = {"not_shopify": 0}

        try:                       # signal handlers only work in the main thread
            signal.signal(signal.SIGINT,  self._stop)
            signal.signal(signal.SIGTERM, self._stop)
        except (ValueError, RuntimeError):
            pass                   # running inside a web worker thread — skip
        init_excel(OUTPUT_EXCEL)

        self.sheets = SheetsUploader(SHEETS_SPREADSHEET_ID,
                                     SHEETS_WORKSHEET_NAME, SHEETS_CREDS_FILE)
        if self.sheets.connect():
            # fold the live sheet's emails into dedup — never re-append
            sheet_emails = self.sheets.existing_emails()
            for em in sheet_emails:
                self.state["emails"].setdefault(hash_key(em), {"source": "sheet"})
            log.info(f"  📥 {len(sheet_emails)} emails already in the sheet (dedup).")

    def _stop(self, sig, frame):
        print()
        log.info("⏸️  Stop — finishing current URL, then saving...")
        self.running = False

    def _flush(self):
        if not self._buf:
            return
        append_to_excel(OUTPUT_EXCEL, self._buf)
        ok = self.sheets.append_rows(self._buf)
        if not ok and self.sheets._ok:
            # Sheet write failed (not a disabled-Sheets run). Roll the dedup
            # entries back so these contacts can be re-collected next run
            # instead of being marked "seen" forever with no row written.
            for rec in self._buf:
                self.state["emails"].pop(hash_key(rec["email"]), None)
            log.warning(f"  ↩️  rolled back {len(self._buf)} dedup entries "
                        f"(sheet write failed) — will retry next run")
        self._buf = []
        self.state["query_counter"]     = self.qgen.counter
        self.state["used_query_hashes"] = list(self.used_hashes)
        self.state["seen_domains"]      = list(self.seen_domains)
        save_state(self.state)
        save_visited(self.visited)

    def _process_url(self, url: str) -> bool:
        if not url or not is_store_url(url):
            return False
        domain = registrable_domain(url)
        dh = hash_key(domain) if domain else ""
        if not domain or dh in self.seen_domains:
            return False
        self.seen_domains.add(dh)             # one shot per domain per DB (stored hashed)

        log.info(f"  🌐 {domain}")
        data = scrape_store(url, self.http, self.visited)

        # redirect may land on an already-seen custom domain
        if data["domain"] != domain:
            dh2 = hash_key(data["domain"]) if data["domain"] else ""
            if dh2 in self.seen_domains:
                return False
            self.seen_domains.add(dh2)

        # ── HARD GATE: must be a Shopify storefront ──────────────
        if not data["shopify"]:
            self.stats["not_shopify"] += 1
            return False

        if _looks_like_junk(data["store"]):
            data["store"] = data["domain"]   # never store listicle titles

        email = data["email"]
        eh = hash_key(email) if email else ""
        if not email or eh in self.state["emails"]:
            return False

        # ── Deliverability + liveness gauntlet ──────────────────
        status, score, reason = self.verifier.verify(
            email, data["domain"], data["page_text"], data["alive"])
        if status == "dead" or score < MIN_SCORE:
            log.info(f"    ❌ {status}/{score} [{reason}]  {email}")
            return False

        # ── Real STORE (not agency/marketplace)? Claude may refine industry ──
        ok_kind, method, industry = self.classifier.validate(
            data["store"], data["url"], data["page_text"], data["industry"])
        if not ok_kind:
            log.info(f"    ❌ not-a-store [{method}]  {email} → {data['store']}")
            return False
        data["industry"] = industry or PROFILE["industry_default"]

        # ── Real mailbox check (optional, rotating keys — runs LAST) ──
        no_credits = self.mailbox.enabled() and self.mailbox.exhausted()
        if no_credits and not self._warned_nocredits:
            log.warning("  ✉️  Mailbox credits exhausted — continuing on FREE gauntlet; "
                        "new rows tagged 'unverified'. Top up to resume mailbox checks.")
            self._warned_nocredits = True

        mbox = self.mailbox.verify(email)   # 'off' if disabled; 'unknown' if no keys left

        if not no_credits:
            if mbox == "undeliverable":
                log.info(f"    ❌ mailbox undeliverable  {email}")
                return False
            if mbox == "catch_all" and MAILBOX_DROP_CATCHALL:
                log.info(f"    ❌ catch-all dropped  {email}")
                return False
            if mbox == "unknown" and MAILBOX_DROP_UNKNOWN and self.mailbox.enabled():
                log.info(f"    ❌ mailbox unverifiable dropped  {email}")
                return False

        if mbox == "deliverable":
            score = min(score + 15, 100)

        if not self.mailbox.enabled():
            mtag = ""
        elif no_credits or mbox == "unknown":
            mtag = "+unverified"
        else:
            mtag = f"+{mbox}"

        theme_s = data["theme"] or "?"
        log.info(f"    ✅ {status}/{score}{mtag}  {email}  →  {data['store']} "
                 f"[{data['industry']} / theme: {theme_s}]")
        rec = {"email": email, "store": data["store"], "url": data["url"],
               "industry": data["industry"], "theme": data["theme"],
               "validated_by": f"{method}+shopify+mx{mtag}",
               "date": datetime.now().strftime("%Y-%m-%d %H:%M")}
        # Committed state (public repo): key by SHA256(email), value carries NO
        # raw email/store/url — only non-identifying metadata for stats.
        # NOTE: 'theme' is intentionally NOT stored here — scraped theme strings
        # sometimes embed the store's domain, which would re-leak PII.
        self.state["emails"][eh] = {"status": status, "score": score,
                                    "industry": data["industry"], "date": rec["date"]}
        self._buf.append(rec)                        # real email → private Excel/Sheets
        if len(self._buf) >= 12:
            self._flush()
        return True

    def _google_phase(self) -> int:
        log.info("─" * 64)
        log.info(f"🔍  GOOGLE PHASE — {MAX_GOOGLE_QUERIES} queries")
        queries = self.qgen.next_batch(MAX_GOOGLE_QUERIES)
        log.info(f"  Generated {len(queries)} queries (counter #{self.qgen.counter})")
        added = 0
        for i, q in enumerate(queries, 1):
            if not self.running:
                break
            if self.google.exhausted():
                log.warning("  ⛔ Google quota exhausted for today — ending "
                            "search phase early (no more keys).")
                break
            log.info(f"\n  [G {i}/{len(queries)}] {q}")
            urls = []
            for page in range(SEARCH_PAGES):
                batch = self.google.search(q, start=1 + page * RESULTS_PER_QUERY)
                urls += batch
                if len(batch) < RESULTS_PER_QUERY:
                    break  # выдача кончилась, вторую страницу не тратим
            log.info(f"    → {len(urls)} URLs ({SEARCH_PAGES} pages max)")
            for url in urls:
                if not self.running:
                    break
                if self._process_url(url):
                    added += 1
                time.sleep(DELAY_URL)
            time.sleep(DELAY_QUERY)
        return added

    def run(self):
        self.state["session_count"] += 1
        before = len(self.state["emails"])
        log.info("=" * 64)
        log.info(f"  Shopify Store Harvester v1.0  │  Session #{self.state['session_count']}")
        log.info(f"  Profile: {PROFILE['name']}   MIN_SCORE={MIN_SCORE}")
        log.info(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        log.info(f"  Contacts in DB    : {before}")
        log.info(f"  Domains processed : {len(self.seen_domains)}")
        log.info(f"  Claude validation : {'ON ✅' if self.classifier.claude.enabled() else 'OFF (rules only)'}")
        log.info(f"  Mailbox API       : {('ON ✅ ('+self.mailbox.provider+')') if self.mailbox.enabled() else 'OFF (free gauntlet only)'}")
        log.info("=" * 64)

        g_new = self._google_phase()
        self._flush()

        after = len(self.state["emails"])
        v, vf = self.classifier.stats, self.verifier.stats
        log.info("=" * 64)
        log.info(f"  Session complete  │  +{g_new} new   total {after}")
        log.info(f"  ── Shopify gate ──")
        log.info(f"    not-Shopify drop : {self.stats['not_shopify']}")
        log.info(f"  ── Deliverability gauntlet ──")
        log.info(f"    live kept        : {vf['live']}")
        log.info(f"    risky (dropped)  : {vf['risky']}")
        log.info(f"    rejected no-MX   : {vf['no_mx']}")
        log.info(f"    rejected mismatch: {vf['mismatch']}")
        log.info(f"    rejected bad-fmt : {vf['bad_format']}")
        log.info(f"    rejected dead-sit: {vf['dead_site']}")
        log.info(f"  ── Store classifier ──")
        log.info(f"    accepted         : {v['accepted']}")
        log.info(f"    rejected         : {v['rejected']}")
        log.info(f"    uncertain dropped: {v['uncertain_dropped']}")
        log.info(f"    claude calls     : {v['claude_calls']}")
        if self.mailbox.enabled():
            mb = self.mailbox.stats
            log.info(f"  ── Mailbox API ({self.mailbox.provider}) ──")
            log.info(f"    deliverable      : {mb['deliverable']}")
            log.info(f"    catch-all        : {mb['catch_all']}")
            log.info(f"    undeliverable    : {mb['undeliverable']}")
            log.info(f"    unknown          : {mb['unknown']}")
            log.info(f"    API credits used : {mb['calls']}")
        log.info(f"  Google API calls  : {self.google.calls}")
        log.info(f"  MX cache size     : {len(self.mx.cache)}")
        log.info(f"  Output (local)    : {OUTPUT_EXCEL}")
        log.info(f"  Google Sheets rows: {self.sheets._sent}")
        log.info("=" * 64)
        self._g_new = g_new

# ═══════════════════════════════════════════════════════════════════
#   🔧  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

def run_once() -> dict:
    """Single batch run — called by run.py / GHA. Returns a summary dict."""
    if not SHEETS_SPREADSHEET_ID or "ЗАМЕНИТЬ" in SHEETS_SPREADSHEET_ID:
        raise RuntimeError(
            "ECOM_SHEET_ID is not set (or still a placeholder). Create the "
            "'Contacts' Google Sheet, share it with the service account, and "
            "set ECOM_SHEET_ID in the workflow before running.")
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    h = EcomHarvester(anthropic_key=key)
    h.run()
    return {
        "parser": "ecom",
        "added": getattr(h, "_g_new", 0),
        "total_in_db": len(h.state["emails"]),
        "sheet_rows_this_run": getattr(h.sheets, "_sent", 0),
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Shopify Store Contact Harvester v1.0 — live stores, deliverable contacts")
    parser.add_argument("--reset", action="store_true", help="clear state, start fresh")
    parser.add_argument("--key", type=str, default="", help="Claude API key (optional)")
    parser.add_argument("--min-score", type=int, default=None, help="quality bar 0-100")
    parser.add_argument("--max-queries", type=int, default=None, help="queries this session")
    parser.add_argument("--mailbox-provider", type=str, default=None,
                        help="millionverifier | reoon (enables real mailbox check)")
    parser.add_argument("--mailbox-keys", type=str, default=None,
                        help="comma-separated API keys (rotated round-robin)")
    args = parser.parse_args()

    if args.min_score is not None:
        MIN_SCORE = args.min_score
    if args.max_queries is not None:
        MAX_GOOGLE_QUERIES = args.max_queries
    if args.mailbox_provider is not None:
        MAILBOX_PROVIDER = args.mailbox_provider.strip()
    if args.mailbox_keys is not None:
        MAILBOX_API_KEYS = [k.strip() for k in args.mailbox_keys.split(",") if k.strip()]
    if args.reset:
        reset_state()

    key = args.key.strip() or ANTHROPIC_API_KEY.strip()
    EcomHarvester(anthropic_key=key).run()
