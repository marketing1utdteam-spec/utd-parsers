#!/usr/bin/env python3
"""
Shopify Content Creator Email Scraper — v4
==========================================

NEW in v4 (vs v3)
-----------------
1. ONE CONTACT PER CREATOR — global dedup across ALL modules and ALL runs.
   A blogger found by the blog module will not reappear via the social module.
   Within a single domain, only the single best-quality email is kept.

2. THEME DEVELOPER BLACKLIST — static list of known Shopify theme makers
   (official Theme Store partners + major third-party makers) PLUS dynamic
   detection: if a domain's homepage shows "buy our theme / our theme
   collection / theme demo" without matching reviewer language → skip.
   These are competitors of UTD BV, not prospects.

3. ACTIVITY / TRAFFIC GATE — each blog/website must score ≥ MIN_ACTIVITY_SCORE
   (default 4 out of 8) based on free proxy signals:
     • Analytics tag present (GA4/UA/Plausible/Fathom/Matomo/Hotjar) → +3
       This is the strongest single signal of a site being seriously operated.
     • RSS feed has posts < 90 days old → +2
     • RSS post < 30 days old (bonus) → +1
     • 10+ pages indexed by Google for "site:domain shopify" → +2
     • 5-9 pages indexed → +1
   Honest note: exact traffic numbers require paid tools (SimilarWeb, Ahrefs).
   The scoring above is a free proxy. A site scoring ≥4 almost certainly has
   real Shopify-focused readership; one scoring <4 is likely dormant or tiny.

   YouTube: channel must have ≥ MIN_YT_SUBSCRIBERS (default 500).

Architecture unchanged from v3 (two-stage discovery → contact pipeline):
  Stage 1: Find WHO regularly covers Shopify (blog domain / YT channel / social profile)
  Stage 2: Go to THEIR /contact or /about page for email — not to the article itself
"""

import os, re, json, time, random, signal, logging, argparse
import xml.etree.ElementTree as ET
from datetime import datetime
from email.utils import parsedate_to_datetime
from urllib.parse import urlparse

import requests
from openpyxl import Workbook, load_workbook

# =========================================================================
# CREDENTIALS
# =========================================================================

API_KEYS = [k.strip() for k in os.environ.get("GOOGLE_API_KEYS", "").split(",") if k.strip()]

GOOGLE_CX_IDS = [c.strip() for c in os.environ.get("GOOGLE_CSE_IDS", "").split(",") if c.strip()]

# =========================================================================
# CONFIG
# =========================================================================

_STATE_DIR            = os.environ.get("STATE_DIR", ".")
os.makedirs(_STATE_DIR, exist_ok=True)
OUTPUT_XLSX           = os.path.join(_STATE_DIR, "shopify_contacts_v4.xlsx")
STATE_FILE            = os.path.join(_STATE_DIR, "scraper_state_v4.json")
PROCESSED_CREATORS    = os.path.join(_STATE_DIR, "processed_creators_v4.json")  # global cross-module creator dedup
LOG_FILE              = os.path.join(_STATE_DIR, "scraper_v4.log")

# --- Google Sheets (live n8n outreach sheet) ------------------------------
# Target: "Copy of IT Companies - Emails", tab Sheet1 (gid 106542427).
# 9 columns: A Email | B Company Name | C Website | D (—) | E (—)
#            F Status | G Date Sent | H Thread ID | I Date Replied
# We append ONLY to A–D (Email, Name, Platform, Profile URL) and never touch
# E–I, so the n8n chain (which fills F–I) keeps working and the column count
# of the document never changes. Append-only — existing rows are preserved.
SHEETS_SPREADSHEET_ID = os.environ.get("INFLUENCER_SHEET_ID", "12IiHIsdibJPRGYNyZfrvdmBDY9OjmsokdmL4GgWg4qQ")
SHEETS_WORKSHEET_GID  = int(os.environ.get("INFLUENCER_SHEET_GID", "106542427"))
SHEETS_CREDS_FILE     = os.environ.get("GOOGLE_CREDS_FILE", os.path.join(_STATE_DIR, "google_credentials.json"))
SHEETS_COLUMNS        = 4   # values appended per row (A–D); MUST stay constant

MAX_CSE_REQUESTS_PER_RUN = 80
MAX_YT_SEARCHES_PER_RUN  = 40
MIN_SHOPIFY_ARTICLES      = 3   # raised from 2 → 3 for higher quality bar
MIN_ACTIVITY_SCORE        = 4   # out of 8, see score_site_activity()
MIN_YT_SUBSCRIBERS        = 500
MAX_EMAILS_PER_CONTACT    = 3   # >3 on a contact page = team directory, skip
PAGE_TIMEOUT              = 12
SEARCH_DELAY              = (1.2, 2.5)
PAGE_DELAY                = (0.8, 2.0)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}

HEADER_ROW = [
    "Email", "Creator / Blog Name", "Platform",
    "Profile / Blog URL", "Email Found At",
    "Shopify Content Links", "Activity Signals", "Date Found",
]

# =========================================================================
# THEME DEVELOPER BLACKLIST (companies that BUILD/SELL Shopify themes)
# =========================================================================
# These are competitors of UTD BV — do NOT contact them.
# Covers: official Shopify Theme Store partners, major marketplace sellers,
# and theme marketplaces.

THEME_CREATOR_DOMAINS = frozenset({
    # Official Shopify Theme Store Partners
    "archetypethemes.co",       # Dawn, Craft, Express, Crave, Colorblock
    "outofthesandbox.com",      # Turbo, Retina, Flex
    "pixelunion.net",           # Empire, Pacific, Startup
    "cleancanvas.co",           # Canopy, Fetch
    "maestrooo.com",            # Symmetry, Streamline
    "fluorescent.co",           # Prestige
    "minionmade.com",           # Motion
    "underground.design",       # Blockshop
    "weareunderground.com",
    "konstructiv.co",           # Impulse
    "boostcommerce.net",
    "halothemes.com",
    "uxtheme.net",
    "ecomify.de",
    "shoptimized.net",
    "fuel-themes.com",
    "mpitheme.com",
    "mpi-themes.com",
    "paperthemes.com",
    "thundertheme.com",
    "streamlinethemes.com",
    "streamline-theme.com",
    "roartheme.com",
    "moosethemes.com",
    "wokiee-theme.com",
    "impulsetheme.com",
    "prestige-theme.com",
    "symmetrytheme.com",
    "minimog-theme.com",
    "boostheme.com",
    "goodthemes.co",
    "velatheme.com",
    "jumplinktheme.com",
    "jumplink.io",
    "debutify.com",
    "looxentertainment.com",
    "getelemental.com",
    "foxtheme.co",
    "foxecom.com",
    "mightythemes.com",
    "uptownthemes.com",
    "labelthemes.com",
    "milestonetheme.com",
    "ecomposer.io",
    "mile.marketing",
    "webibazaar.com",
    "webinopoly.com",
    "themezaa.com",
    "ellipsis.studio",
    # More official Shopify Theme Store partners
    "krownthemes.com",          # Local, Loft, Symmetry-family
    "krownthemes.co",
    "groupthought.com",         # Pipeline
    "switchthemes.co",          # Stiletto, Vessel
    "weareeight.com",           # Be Yours, Yuva, Studio
    "shrinetheme.com",
    "shrineapp.io",
    "presidiocreative.com",
    "troopthemes.com",
    "eightthemes.com",
    "rightoolsthemes.com",
    "kaktusthemes.com",
    "zerogrocery.studio",
    "fastorthemes.com",
    "qartzthemes.com",
    "themepublishers.com",
    # Theme Marketplaces
    "themeforest.net",
    "templatemonster.com",
    "envato.com",
    "creative-tim.com",
    # Own company — never contact yourself
    "utdweb.team",
})

# Text signals that a site BUILDS/SELLS themes (not reviews them)
THEME_CREATOR_SIGNALS = (
    "buy our theme", "purchase our theme", "download our shopify theme",
    "our premium shopify theme", "shop our themes", "our theme collection",
    "we create shopify themes", "our shopify themes", "theme by our team",
    "we built this theme", "theme developed by us", "purchase this theme",
    "our themes start at $", "get our theme", "theme download",
    "built by our team", "our own theme",
)

REVIEWER_SIGNALS = (
    "theme review", "i tested", "best shopify themes of",
    "theme comparison", "pros and cons", "our verdict", "star rating",
    "shopify theme review", "reviewed by",
)


def is_theme_creator(domain, html=""):
    """
    Returns True if this domain BUILDS or SELLS Shopify themes.
    First checks static blacklist, then scans homepage HTML for creator signals.
    """
    d = domain.lower().lstrip("www.")
    if any(d == bl or d.endswith("." + bl) for bl in THEME_CREATOR_DOMAINS):
        return True
    if html:
        text = html.lower()
        creator_hits   = sum(1 for s in THEME_CREATOR_SIGNALS if s in text)
        reviewer_hits  = sum(1 for s in REVIEWER_SIGNALS     if s in text)
        if creator_hits >= 2 and creator_hits > reviewer_hits:
            logging.info(f"  ✗ Dynamic detection: {domain} looks like theme creator "
                         f"(creator_hits={creator_hits}, reviewer_hits={reviewer_hits})")
            return True
    return False


# =========================================================================
# ACTIVITY SCORING (free proxy for ~1000+ monthly visitors)
# =========================================================================

ANALYTICS_RE = re.compile(
    r'G-[A-Z0-9]{8,12}'          # GA4
    r'|UA-\d{4,9}-\d{1,4}'       # Universal Analytics
    r'|gtm\.js'                   # Google Tag Manager
    r'|gtag\s*\('                 # gtag function
    r'|plausible\.io'
    r'|fathom'
    r'|matomo'
    r'|heap\.io'
    r'|hotjar'
    r'|clarity\.ms'               # Microsoft Clarity
    r'|_paq'                      # Matomo/Piwik
)

RSS_PATHS = [
    "/feed", "/rss.xml", "/feed.xml", "/rss", "/atom.xml",
    "/feeds/posts/default", "/blog/feed", "/blog/rss.xml",
    "/blog/feed.xml", "/en/feed", "/news/feed",
]


def has_analytics(html):
    return bool(ANALYTICS_RE.search(html or ""))


def check_rss_activity(domain):
    """
    Fetches the site's RSS/Atom feed (tries common paths).
    Returns (has_recent_post: bool, post_count: int, days_since_last: int).
    'Recent' = last post within 90 days.
    """
    for path in RSS_PATHS:
        url = f"https://{domain}{path}"
        try:
            r = requests.get(url, headers=HEADERS, timeout=8)
            body = r.text[:80_000]
            if r.status_code != 200:
                continue
            if not any(tag in body[:2000] for tag in ("<rss", "<feed", "<channel", "<?xml")):
                continue
            tree = ET.fromstring(body)
            dates = []
            # RSS 2.0
            for item in tree.iter("item"):
                pd = item.find("pubDate")
                if pd is not None and pd.text:
                    try:
                        dates.append(
                            parsedate_to_datetime(pd.text.strip()).replace(tzinfo=None)
                        )
                    except Exception:
                        pass
            # Atom 1.0
            for entry in tree.iter("{http://www.w3.org/2005/Atom}entry"):
                upd = entry.find("{http://www.w3.org/2005/Atom}updated")
                if upd is not None and upd.text:
                    try:
                        dates.append(datetime.fromisoformat(upd.text[:19]))
                    except Exception:
                        pass
            if dates:
                dates.sort(reverse=True)
                days_ago = (datetime.now() - dates[0]).days
                return days_ago < 90, len(dates), days_ago
        except Exception:
            pass
    return False, 0, 999


def score_site_activity(domain, homepage_html, cse_indexed_count):
    """
    Returns (score: int, signals: list[str]).
    Needs MIN_ACTIVITY_SCORE (4) to pass.

    Points:
      Analytics tag:          +3  (strongest free signal of serious operation)
      RSS with post <90d:     +2
      RSS with post <30d:     +1 bonus
      CSE indexed ≥10:        +2
      CSE indexed 5-9:        +1
    Max: 8
    """
    score, signals = 0, []

    if has_analytics(homepage_html or ""):
        score += 3
        signals.append("analytics")

    has_recent, post_count, days_ago = check_rss_activity(domain)
    if has_recent:
        score += 2
        signals.append(f"rss:{post_count}posts:{days_ago}d")
        if days_ago < 30:
            score += 1
            signals.append("active<30d")

    if cse_indexed_count >= 10:
        score += 2
        signals.append(f"indexed≥10")
    elif cse_indexed_count >= 5:
        score += 1
        signals.append(f"indexed:{cse_indexed_count}")

    return score, signals


# =========================================================================
# KEY ROTATOR
# =========================================================================

class KeyRotator:
    def __init__(self, keys):
        self.keys = keys
        self.idx = 0
        self.exhausted = set()

    def current(self):
        return self.keys[self.idx]

    def rotate(self, reason="quota"):
        logging.warning(f"Key idx {self.idx} exhausted ({reason}), rotating.")
        self.exhausted.add(self.idx)
        self.idx = (self.idx + 1) % len(self.keys)

    def all_exhausted(self):
        return len(self.exhausted) >= len(self.keys)

    def cx(self):
        return GOOGLE_CX_IDS[self.idx % len(GOOGLE_CX_IDS)]


cse_rotator = KeyRotator(API_KEYS)
yt_rotator  = KeyRotator(API_KEYS)

# =========================================================================
# EMAIL EXTRACTION + VALIDATION
# =========================================================================

EMAIL_RE    = re.compile(r"[a-zA-Z0-9_.+-]{2,40}@[a-zA-Z0-9-]{2,40}(?:\.[a-zA-Z0-9-]{1,20})*\.[a-zA-Z]{2,20}")
CF_EMAIL_RE = re.compile(r'data-cfemail="([a-f0-9]+)"')
AT_DOT_RE   = re.compile(
    r"([a-zA-Z0-9_.+-]{2,30})\s*[\[\(]?\s*(?:at|AT)\s*[\]\)]?\s*"
    r"([a-zA-Z0-9-]{2,30})\s*[\[\(]?\s*(?:dot|DOT)\s*[\]\)]?\s*([a-zA-Z]{2,20})",
)

BLOCKED_LOCALS = {
    "noreply", "no-reply", "donotreply", "do-not-reply", "webmaster",
    "postmaster", "abuse", "unsubscribe", "mailer-daemon", "bounce",
    "bounces", "billing", "legal", "privacy", "security", "admin",
    "administrator", "test", "testing",
}

BLOCKED_DOMAINS = {
    "example.com", "yourdomain.com", "domain.com", "email.com", "schema.org",
    "sentry.io", "wixpress.com", "godaddy.com", "partnerstack.com",
    "stamped.io", "gempages.net", "globo.io", "channels.app",
    "shopify.com", "myshopify.com", "youtube.com", "instagram.com",
    "tiktok.com", "facebook.com", "twitter.com", "linkedin.com",
    "apple.com", "google.com",
}

PERSONAL_DOMAINS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com",
    "icloud.com", "protonmail.com", "pm.me", "yandex.com",
}

EMAIL_QUALITY_RE = re.compile(r"^[a-z]{3,15}$")  # pure lowercase word = likely a name


def decode_cf_email(enc):
    try:
        r = int(enc[:2], 16)
        return "".join(chr(int(enc[i:i+2], 16) ^ r) for i in range(2, len(enc), 2))
    except Exception:
        return None


def validate_email(raw):
    e = re.sub(r'[.,;:)\]>"\s]+$', '', raw.strip()).lower()
    if not EMAIL_RE.fullmatch(e):
        return None
    local, domain = e.split("@", 1)
    if local in BLOCKED_LOCALS:
        return None
    if re.match(r'^u[0-9a-f]{3,5}$', local):  # JS escape artifact
        return None
    if domain in PERSONAL_DOMAINS:
        return e  # personal gmail etc. OK
    if any(domain == d or domain.endswith("." + d) for d in BLOCKED_DOMAINS):
        return None
    return e


def score_email(email):
    """Higher = more personal/direct. Used to pick the best email from a contact page."""
    local, domain = email.split("@", 1)
    if "." in local or "_" in local:
        return 4   # firstname.lastname@ or first_last@
    if EMAIL_QUALITY_RE.match(local) and local not in ("info", "hello", "contact", "team"):
        return 3   # single-word that's probably a name
    if local in ("hello", "hi", "team", "editor", "writer", "creator"):
        return 2
    if local in ("info", "contact", "media", "press", "advertise"):
        return 1
    return 0


def best_email(emails):
    """Pick the single highest-quality email from a set."""
    if not emails:
        return None
    return max(emails, key=score_email)


def extract_emails_from_html(html):
    candidates = set()
    for enc in CF_EMAIL_RE.findall(html):
        d = decode_cf_email(enc)
        if d:
            candidates.add(d)
    candidates.update(EMAIL_RE.findall(html))
    for m in AT_DOT_RE.findall(html):
        candidates.add(f"{m[0]}@{m[1]}.{m[2]}")
    return {v for c in candidates for v in [validate_email(c)] if v}


# =========================================================================
# HTTP HELPERS
# =========================================================================

def fetch_html(url, timeout=PAGE_TIMEOUT):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and "text/html" in r.headers.get("Content-Type", ""):
            return r.text
    except Exception:
        pass
    return None


def cse_search(query, start=1):
    """
    Returns (items: list, error: str|None, total_results: int).
    total_results is Google's estimated indexed count for the query.
    """
    key = cse_rotator.current()
    cx  = cse_rotator.cx()
    try:
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key": key, "cx": cx, "q": query, "start": start, "num": 10},
            timeout=15,
        )
        if r.status_code == 200:
            data  = r.json()
            total = int(data.get("searchInformation", {}).get("totalResults", "0"))
            return data.get("items", []), None, total
        if r.status_code in (403, 429):
            cse_rotator.rotate("quota/rate")
            return [], "rotated", 0
        return [], f"http_{r.status_code}", 0
    except Exception as exc:
        return [], f"network:{exc}", 0


def yt_api(endpoint, params):
    base = "https://www.googleapis.com/youtube/v3/"
    for _ in range(len(API_KEYS)):
        p = dict(params)
        p["key"] = yt_rotator.current()
        try:
            r = requests.get(base + endpoint, params=p, timeout=15)
            if r.status_code == 200:
                return r.json(), None
            if r.status_code in (403, 429):
                yt_rotator.rotate("quota/rate")
                continue
            return None, f"http_{r.status_code}"
        except Exception as exc:
            return None, f"network:{exc}"
    return None, "all_keys_exhausted"


# =========================================================================
# CONTACT PAGE FINDER
# =========================================================================

CONTACT_PATHS = [
    "/contact", "/contact-us", "/about", "/about-us",
    "/contact.html", "/about.html", "/contact/",
    "/advertise", "/advertise-with-us", "/work-with-me",
    "/work-with-us", "/collaborate", "/partner",
    "/press", "/media", "/hire-me", "/reach-out",
]


def find_best_contact_email(base_url):
    """
    Tries CONTACT_PATHS, returns (best_email_str, source_url) or (None, None).
    Enforces MAX_EMAILS_PER_CONTACT per page (>3 = team directory, skip).
    Returns the single best-scored email per page — never multiple.
    """
    parsed = urlparse(base_url)
    root   = f"{parsed.scheme}://{parsed.netloc}"

    for path in CONTACT_PATHS:
        html = fetch_html(root + path)
        if not html:
            continue
        emails = extract_emails_from_html(html)
        if 0 < len(emails) <= MAX_EMAILS_PER_CONTACT:
            chosen = best_email(emails)
            if chosen:
                return chosen, root + path
        time.sleep(random.uniform(0.3, 0.6))

    # Fallback: homepage
    html = fetch_html(root)
    if html:
        emails = extract_emails_from_html(html)
        if 0 < len(emails) <= MAX_EMAILS_PER_CONTACT:
            chosen = best_email(emails)
            if chosen:
                return chosen, root

    return None, None


# =========================================================================
# URL / DOMAIN HELPERS
# =========================================================================

EXCLUDED_DOMAIN_TOKENS = (
    "shopify.com", "myshopify.com", "apps.apple.com", "play.google.com",
    "amazon.com", "reddit.com", "pinterest.com", "facebook.com",
    "twitter.com", "linkedin.com", "youtube.com", "tiktok.com",
    "instagram.com", "vimeo.com", "coursera.org", "udemy.com",
    "quora.com", "wikipedia.org", "github.com",
)

EXCLUDED_PATH_TOKENS = (
    "/hc/", "/help-center/", "/support/", "/legal/", "/terms",
    "/privacy", "/sitemap", "/tag/", "/wp-admin", "/wp-login",
    "/checkout", "/cart", "/products/", "/collections/",
    "/careers", "/job", "/press-release", "/forum", "/community/",
)


def domain_of(url):
    return urlparse(url).netloc.lower().lstrip("www.")


def is_excluded_domain(url):
    d = domain_of(url)
    return any(t in d for t in EXCLUDED_DOMAIN_TOKENS)


def is_excluded_path(url):
    p = urlparse(url).path.lower()
    return any(t in p for t in EXCLUDED_PATH_TOKENS)


# =========================================================================
# CREATOR REGISTRY — cross-module, cross-run dedup
# =========================================================================

def creator_key(platform, identifier):
    """Stable key for a unique creator entity."""
    return f"{platform.lower()}:{identifier.lower().lstrip('www.')}"


# =========================================================================
# MODULE A — BLOG / WEBSITE FINDER
# =========================================================================

BLOG_DISCOVERY_QUERIES = [
    "shopify theme review blog",
    "best shopify themes review site",
    "shopify theme comparison blog",
    "shopify theme tutorial blog post",
    "shopify ecommerce design review",
    "shopify theme honest review site",
    '"shopify theme" review -shopify.com intitle:review',
    "shopify theme reviewer blog",
    "shopify store design blog review",
    "I tested shopify theme review blog",
    "shopify theme walkthrough review",
    "new shopify theme 2025 review blog",
    "shopify theme review newsletter",
    "shopify theme expert review",
    "top shopify themes 2026 review article",
    "shopify theme review podcast blog",
    '"shopify theme" review site:substack.com',
    '"shopify theme" review site:medium.com',
    '"shopify" theme review "contact us"',
    '"shopify themes" comparison "write for us"',
]


def run_blog_module(state, seen_creators, seen_emails, max_requests, stop):
    new_rows = []
    req_used = state.get("blog_cse_used", 0)
    q_idx    = state.get("blog_query_idx", 0)
    start    = state.get("blog_start", 1)

    logging.info(f"=== MODULE A: BLOGS === budget={max_requests}")

    while req_used < max_requests and not stop["flag"]:
        if q_idx >= len(BLOG_DISCOVERY_QUERIES):
            logging.info("Blog queries exhausted.")
            break

        query = BLOG_DISCOVERY_QUERIES[q_idx]
        items, err, _ = cse_search(query, start)
        req_used += 1
        time.sleep(random.uniform(*SEARCH_DELAY))

        if cse_rotator.all_exhausted():
            logging.error("All CSE keys exhausted.")
            break

        logging.info(f"Blog [{req_used}/{max_requests}] '{query}' start={start} → {len(items)} items")

        domains_this_batch = {}
        for item in items:
            link = item.get("link", "")
            if not link or is_excluded_domain(link) or is_excluded_path(link):
                continue
            d = domain_of(link)
            if d and d not in domains_this_batch:
                domains_this_batch[d] = (link, item.get("title", ""))

        for domain, (sample_url, title) in domains_this_batch.items():
            if stop["flag"] or req_used >= max_requests:
                break

            ck = creator_key("blog", domain)
            if ck in seen_creators:
                logging.info(f"  ⟳ {domain} already in creator registry, skip")
                continue

            # ---- Verify: regular Shopify publisher? ----
            items2, err2, total_indexed = cse_search(f"site:{domain} shopify theme")
            req_used += 1
            time.sleep(random.uniform(*SEARCH_DELAY))

            article_urls = [i.get("link", "") for i in items2
                            if i.get("link") and not is_excluded_path(i["link"])]

            if len(article_urls) < MIN_SHOPIFY_ARTICLES:
                logging.info(f"  ✗ {domain}: only {len(article_urls)} Shopify articles (<{MIN_SHOPIFY_ARTICLES})")
                seen_creators[ck] = "skipped:low_articles"
                continue

            # ---- Theme creator check ----
            homepage_html = fetch_html(f"https://{domain}")
            time.sleep(random.uniform(*PAGE_DELAY))

            if is_theme_creator(domain, homepage_html or ""):
                logging.info(f"  ✗ {domain}: identified as Shopify theme creator/seller")
                seen_creators[ck] = "skipped:theme_creator"
                continue

            # ---- Activity / traffic gate ----
            score, signals = score_site_activity(domain, homepage_html, total_indexed)
            logging.info(f"  Activity: {domain} score={score}/8 signals={signals}")
            if score < MIN_ACTIVITY_SCORE:
                logging.info(f"  ✗ {domain}: activity score {score} < {MIN_ACTIVITY_SCORE}, skip")
                seen_creators[ck] = f"skipped:low_activity:{score}"
                continue

            # ---- Find contact email ----
            email, email_source = find_best_contact_email(sample_url)
            time.sleep(random.uniform(*PAGE_DELAY))

            if not email:
                logging.info(f"  ✗ {domain}: no contact email found")
                seen_creators[ck] = "skipped:no_email"
                continue

            if email in seen_emails:
                seen_creators[ck] = "skipped:dup_email"
                continue

            seen_emails.add(email)
            seen_creators[ck] = email
            signals_str = ", ".join(signals)
            new_rows.append({
                "email":        email,
                "name":         title.split("|")[0].split(" - ")[0].strip() or domain,
                "platform":     "Blog/Website",
                "profile_url":  f"https://{domain}",
                "email_source": email_source or f"https://{domain}",
                "content_links": " | ".join(article_urls[:3]),
                "signals":      signals_str,
                "date":         datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            logging.info(f"  ✉ {email} | {domain} | score={score} | {signals_str}")

        start += 10
        if start > 91:
            start = 1
            q_idx += 1

        state.update(blog_cse_used=req_used, blog_query_idx=q_idx, blog_start=start)

    return new_rows, req_used


# =========================================================================
# MODULE B — YOUTUBE CHANNEL FINDER
# =========================================================================

YT_SEARCH_QUERIES = [
    "shopify theme review",
    "best shopify themes",
    "shopify theme comparison",
    "shopify theme tutorial",
    "shopify store design review",
    "shopify ecommerce theme walkthrough",
    "shopify theme honest review",
    "shopify theme 2025 2026 review",
    "top shopify themes explained",
    "shopify theme setup guide",
]


def extract_external_links(text):
    return re.findall(r'https?://[^\s\)\]>"\',]+', text or "")


def extract_email_from_text(text):
    return {v for c in EMAIL_RE.findall(text or "") for v in [validate_email(c)] if v}


def run_youtube_module(state, seen_creators, seen_emails, max_searches, stop):
    new_rows = []
    srch_used  = state.get("yt_searches_used", 0)
    q_idx      = state.get("yt_query_idx", 0)
    page_token = state.get("yt_page_token", "")

    logging.info(f"=== MODULE B: YOUTUBE === budget={max_searches}")

    while srch_used < max_searches and not stop["flag"]:
        if q_idx >= len(YT_SEARCH_QUERIES):
            logging.info("YT queries exhausted.")
            break

        query  = YT_SEARCH_QUERIES[q_idx]
        params = {"part": "snippet", "q": query, "type": "channel",
                  "maxResults": 25, "relevanceLanguage": "en"}
        if page_token:
            params["pageToken"] = page_token

        data, err = yt_api("search", params)
        srch_used += 1
        time.sleep(random.uniform(*SEARCH_DELAY))

        if err:
            logging.warning(f"YT search error: {err}")
            if "exhausted" in (err or ""):
                break
            q_idx += 1
            page_token = ""
            continue

        items      = data.get("items", [])
        next_token = data.get("nextPageToken", "")
        logging.info(f"YT [{srch_used}/{max_searches}] '{query}' → {len(items)} channels")

        new_ch_ids = [
            i["id"]["channelId"] for i in items
            if i.get("id", {}).get("kind") == "youtube#channel"
            and creator_key("yt", i["id"]["channelId"]) not in seen_creators
        ]

        if new_ch_ids:
            ch_data, _ = yt_api("channels", {
                "part": "snippet,brandingSettings,statistics",
                "id": ",".join(new_ch_ids),
                "maxResults": 50,
            })
            time.sleep(random.uniform(0.5, 1.0))

            for ch in (ch_data or {}).get("items", []):
                ch_id  = ch["id"]
                ck     = creator_key("yt", ch_id)

                if ck in seen_creators:
                    continue

                snippet  = ch.get("snippet", {})
                branding = ch.get("brandingSettings", {}).get("channel", {})
                stats    = ch.get("statistics", {})
                name     = snippet.get("title", "")
                desc     = snippet.get("description", "")
                custom   = snippet.get("customUrl", "")
                subs     = int(stats.get("subscriberCount", "0") or "0")
                yt_url   = f"https://www.youtube.com/{'@' + custom.lstrip('@') if custom else 'channel/' + ch_id}"

                # Subscriber threshold
                if subs < MIN_YT_SUBSCRIBERS:
                    logging.info(f"  ✗ {name}: {subs} subscribers < {MIN_YT_SUBSCRIBERS}")
                    seen_creators[ck] = f"skipped:low_subs:{subs}"
                    continue

                # Theme creator check (channel name / description)
                if is_theme_creator("", desc.lower()):
                    logging.info(f"  ✗ {name}: looks like a theme maker channel")
                    seen_creators[ck] = "skipped:theme_creator"
                    continue

                # 1) Email directly in description
                desc_emails = extract_email_from_text(desc)

                # 2) External site → /contact page
                ext_email, contact_src = None, None
                for link in extract_external_links(desc)[:3]:
                    if is_excluded_domain(link):
                        continue
                    ext_domain = domain_of(link)
                    if ext_domain in seen_creators.values():
                        continue
                    e, src = find_best_contact_email(link)
                    if e:
                        ext_email, contact_src = e, src
                        break
                    time.sleep(random.uniform(*PAGE_DELAY))

                all_emails = desc_emails | ({ext_email} if ext_email else set())

                if not all_emails:
                    logging.info(f"  ✗ {name}: no email found")
                    seen_creators[ck] = "skipped:no_email"
                    continue

                chosen = best_email(all_emails)
                if not chosen or chosen in seen_emails:
                    seen_creators[ck] = "skipped:dup_email"
                    continue

                # Collect proof videos
                vid_data, _ = yt_api("search", {
                    "part": "snippet", "channelId": ch_id,
                    "q": "shopify theme", "type": "video", "maxResults": 3,
                })
                video_links = [
                    f"https://www.youtube.com/watch?v={v['id']['videoId']}"
                    for v in (vid_data or {}).get("items", [])
                    if v.get("id", {}).get("videoId")
                ]

                seen_emails.add(chosen)
                seen_creators[ck] = chosen
                new_rows.append({
                    "email":        chosen,
                    "name":         name,
                    "platform":     "YouTube",
                    "profile_url":  yt_url,
                    "email_source": contact_src or yt_url,
                    "content_links": " | ".join(video_links[:3]),
                    "signals":      f"subs:{subs}",
                    "date":         datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
                logging.info(f"  ✉ YT: {chosen} | {name} | subs={subs}")

        page_token = next_token
        if not page_token:
            q_idx += 1
            page_token = ""

        state.update(yt_searches_used=srch_used, yt_query_idx=q_idx, yt_page_token=page_token)

    return new_rows, srch_used


# =========================================================================
# MODULE C — SOCIAL MEDIA PROFILE FINDER
# =========================================================================

SOCIAL_QUERIES = [
    ('site:instagram.com "shopify theme" review creator', "Instagram"),
    ('site:instagram.com "shopify themes" ecommerce tutorial', "Instagram"),
    ('site:instagram.com shopify store design blog creator', "Instagram"),
    ('site:tiktok.com "shopify theme" review', "TikTok"),
    ('site:tiktok.com shopify ecommerce tutorial creator', "TikTok"),
    ('site:tiktok.com "shopify" theme blogger', "TikTok"),
    ('site:linkedin.com/in/ "shopify theme" content creator blogger', "LinkedIn"),
    ('site:linkedin.com/in/ shopify "content creator" ecommerce themes', "LinkedIn"),
    ('site:twitter.com "shopify theme" review creator', "Twitter/X"),
    ('site:twitter.com shopify ecommerce "content creator"', "Twitter/X"),
    ('site:x.com "shopify theme" review creator', "Twitter/X"),
    ('site:x.com shopify ecommerce design creator', "Twitter/X"),
    ('site:substack.com shopify theme review newsletter', "Substack"),
    ('site:substack.com shopify ecommerce design newsletter', "Substack"),
    ('site:pinterest.com shopify theme design creator', "Pinterest"),
    ('site:threads.net shopify theme ecommerce creator', "Threads"),
    ('site:t.me shopify theme review channel', "Telegram"),
    ('site:t.me shopify ecommerce news', "Telegram"),
]

EXT_LINK_RE = re.compile(
    r'https?://(?!(?:www\.)?(?:instagram|tiktok|linkedin|facebook|twitter|x|t\.me|'
    r'youtube|pinterest|threads|substack)\.com)'
    r'[^\s\)\]>"\',]{10,}'
)


def extract_website_from_profile(html):
    for m in EXT_LINK_RE.findall(html or ""):
        m = m.rstrip("/.,;")
        if not is_excluded_domain(m) and len(m) > 12:
            return m
    return None


def run_social_module(state, seen_creators, seen_emails, max_requests, stop):
    new_rows = []
    req_used = state.get("social_cse_used", 0)
    q_idx    = state.get("social_query_idx", 0)

    logging.info(f"=== MODULE C: SOCIAL === budget={max_requests}")

    while req_used < max_requests and not stop["flag"]:
        if q_idx >= len(SOCIAL_QUERIES):
            logging.info("Social queries exhausted.")
            break

        query, platform = SOCIAL_QUERIES[q_idx]
        items, err, _   = cse_search(query)
        req_used += 1
        time.sleep(random.uniform(*SEARCH_DELAY))

        if cse_rotator.all_exhausted():
            break

        logging.info(f"Social [{req_used}/{max_requests}] [{platform}] → {len(items)} profiles")

        for item in items:
            if stop["flag"]:
                break
            profile_url = item.get("link", "")
            if not profile_url:
                continue

            ck = creator_key(platform, profile_url)
            if ck in seen_creators:
                continue

            profile_name = item.get("title", "").split("|")[0].strip()
            html = fetch_html(profile_url)
            time.sleep(random.uniform(*PAGE_DELAY))

            # Direct email in profile page (some TikTok/Substack pages show it)
            direct = extract_emails_from_html(html) if html else set()
            direct = {e for e in direct if validate_email(e)}

            # External website from profile → /contact
            external_site = extract_website_from_profile(html) if html else None
            ext_email, contact_src = None, None

            if external_site:
                ext_d = domain_of(external_site)
                ext_ck = creator_key("blog", ext_d)
                if ext_ck not in seen_creators and not is_theme_creator(ext_d):
                    ext_email, contact_src = find_best_contact_email(external_site)
                    time.sleep(random.uniform(*PAGE_DELAY))
                    if ext_email:
                        seen_creators[ext_ck] = ext_email

            all_emails = direct | ({ext_email} if ext_email else set())
            if not all_emails:
                seen_creators[ck] = "skipped:no_email"
                continue

            chosen = best_email(all_emails)
            if not chosen or chosen in seen_emails:
                seen_creators[ck] = "skipped:dup_email"
                continue

            seen_emails.add(chosen)
            seen_creators[ck] = chosen
            new_rows.append({
                "email":        chosen,
                "name":         profile_name,
                "platform":     platform,
                "profile_url":  profile_url,
                "email_source": contact_src or profile_url,
                "content_links": profile_url,
                "signals":      f"social_profile",
                "date":         datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            logging.info(f"  ✉ {platform}: {chosen} | {profile_name}")

        q_idx += 1
        state.update(social_cse_used=req_used, social_query_idx=q_idx)

    return new_rows, req_used


# =========================================================================
# EXCEL + STATE PERSISTENCE
# =========================================================================

def load_existing_emails(path):
    seen = set()
    if os.path.exists(path):
        wb = load_workbook(path)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                seen.add(str(row[0]).lower())
    return seen


def append_to_excel(path, rows):
    if not rows:
        return
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADER_ROW)
    for r in rows:
        ws.append([r["email"], r["name"], r["platform"], r["profile_url"],
                   r["email_source"], r["content_links"], r["signals"], r["date"]])
    wb.save(path)


def load_json(path, default):
    try:
        if os.path.exists(path):
            with open(path) as f:
                return json.load(f)
    except Exception:
        pass
    return default


def save_json(path, data):
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


# =========================================================================
# GOOGLE SHEETS SYNC  (append-only, n8n-safe — never changes column count)
# =========================================================================

class SheetsSync:
    """
    Appends new contacts to the live n8n Google Sheet WITHOUT changing its
    column count and WITHOUT overwriting existing rows.

    Writes ONLY columns A–D (Email, Name, Platform, Profile URL).
    Columns E–I — including n8n's Status / Date Sent / Thread ID /
    Date Replied — are never touched, so the automation keeps working.
    """

    def __init__(self, spreadsheet_id, gid, creds_file):
        self.spreadsheet_id = spreadsheet_id
        self.gid            = gid
        self.creds_file     = creds_file
        self._ws            = None
        self.ok             = False
        self.sent           = 0

    def connect(self):
        if not os.path.exists(self.creds_file):
            logging.warning(f"Sheets: creds file not found ({self.creds_file}); local xlsx only.")
            return False
        try:
            import gspread
        except ImportError:
            logging.error("Sheets: gspread not installed "
                          "(pip install gspread google-auth); local xlsx only.")
            return False
        try:
            gc = gspread.service_account(filename=self.creds_file)
            ss = gc.open_by_key(self.spreadsheet_id)
            self._ws = ss.get_worksheet_by_id(self.gid) or ss.get_worksheet(0)
            if self._ws is None:
                logging.error("Sheets: worksheet not found; local xlsx only.")
                return False
            self.ok = True
            logging.info(f"Sheets: connected → tab '{self._ws.title}' in '{ss.title}'")
            return True
        except Exception as e:
            logging.error(f"Sheets: connect failed ({e}); local xlsx only.")
            return False

    def existing_emails(self):
        """Emails already in column A — so we never duplicate or overwrite."""
        if not self.ok:
            return set()
        try:
            col = self._ws.col_values(1)
            return {str(v).strip().lower() for v in col[1:] if v and "@" in str(v)}
        except Exception as e:
            logging.warning(f"Sheets: could not read existing emails ({e})")
            return set()

    def append(self, rows):
        """
        Append A–D only (exactly SHEETS_COLUMNS values per row), append-only.

        IMPORTANT: writes with a direct update() at the true last row + 1,
        NOT append_rows(). The sheet contains a blank row in the middle, and
        append_rows() would insert at that gap and shift every following row
        (corrupting the n8n columns). update() writes the block at the real
        end without inserting or shifting anything — gap-proof and safe.
        """
        if not self.ok or not rows:
            return False
        values = [[r["email"], r["name"], r["platform"], r["profile_url"]]
                  for r in rows]
        for attempt in range(1, 4):
            try:
                next_row = len(self._ws.get_all_values()) + 1   # true end
                self._ws.update(
                    range_name=f"A{next_row}",
                    values=values,
                    value_input_option="USER_ENTERED",
                )
                self.sent += len(values)
                logging.info(f"Sheets: +{len(values)} rows written at A{next_row} "
                             f"(session total {self.sent})")
                return True
            except Exception as e:
                err = str(e).lower()
                if any(k in err for k in ("429", "quota", "rate")):
                    time.sleep(20 * attempt)
                elif "401" in err or "403" in err:
                    logging.error(f"Sheets: auth/permission error ({e}); disabling sheet sync.")
                    self.ok = False
                    return False
                else:
                    time.sleep(5 * attempt)
        logging.error("Sheets: append failed after retries; rows kept in local xlsx.")
        return False


def checkpoint(state, seen_creators, new_rows, seen_emails, output_path, sheets=None):
    append_to_excel(output_path, new_rows)
    if sheets is not None:
        sheets.append(new_rows)          # push to Google Sheets (append-only)
    save_json(STATE_FILE, state)
    save_json(PROCESSED_CREATORS, seen_creators)
    logging.info(f"✓ Checkpoint. Emails collected: {len(seen_emails)}")


# =========================================================================
# MAIN
# =========================================================================

def main(cse_budget, yt_budget, output_path):
    setup_logging()

    state        = load_json(STATE_FILE, {})
    seen_creators = load_json(PROCESSED_CREATORS, {})   # {creator_key: email_or_status}
    seen_emails  = load_existing_emails(output_path)

    # Google Sheets — connect and fold its existing emails into the dedup set
    # so we never append a contact that is already in the live n8n sheet.
    sheets = SheetsSync(SHEETS_SPREADSHEET_ID, SHEETS_WORKSHEET_GID, SHEETS_CREDS_FILE)
    if sheets.connect():
        sheet_emails = sheets.existing_emails()
        seen_emails |= sheet_emails
        logging.info(f"Startup: {len(sheet_emails)} emails already in Google Sheet.")

    logging.info(
        f"Startup: {len(seen_emails)} emails known (xlsx+sheet), "
        f"{len(seen_creators)} creator keys in registry."
    )

    stop = {"flag": False}
    def handle_stop(sig, frame):
        logging.info("Stop signal — saving and exiting...")
        stop["flag"] = True
    try:                       # signal handlers only work in the main thread
        signal.signal(signal.SIGINT,  handle_stop)
        signal.signal(signal.SIGTERM, handle_stop)
    except (ValueError, RuntimeError):
        pass                   # running inside a web worker thread — skip

    rows_a, rows_b, rows_c = [], [], []

    # Module A — Blogs
    blog_budget = cse_budget // 2
    rows_a, _   = run_blog_module(state, seen_creators, seen_emails, blog_budget, stop)
    checkpoint(state, seen_creators, rows_a, seen_emails, output_path, sheets)

    # Module B — YouTube
    if not stop["flag"]:
        rows_b, _ = run_youtube_module(state, seen_creators, seen_emails, yt_budget, stop)
        checkpoint(state, seen_creators, rows_b, seen_emails, output_path, sheets)

    # Module C — Social
    if not stop["flag"]:
        social_budget = cse_budget - blog_budget
        rows_c, _     = run_social_module(state, seen_creators, seen_emails,
                                          social_budget, stop)
        checkpoint(state, seen_creators, rows_c, seen_emails, output_path, sheets)

    added = len(rows_a) + len(rows_b) + len(rows_c)
    logging.info(f"Run complete. Output: {output_path} | Sheet rows added: {sheets.sent}")
    return {"parser": "influencer", "added": added,
            "blog": len(rows_a), "youtube": len(rows_b), "social": len(rows_c),
            "sheet_rows_this_run": sheets.sent}


def run_once() -> dict:
    """Single batch run — called by the web service. Returns a summary dict."""
    cse = int(os.environ.get("CSE_BUDGET", "20"))
    yt  = int(os.environ.get("YT_BUDGET", "10"))
    return main(cse, yt, OUTPUT_XLSX)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Shopify creator email scraper v4")
    parser.add_argument("--cse-budget", type=int, default=MAX_CSE_REQUESTS_PER_RUN)
    parser.add_argument("--yt-budget",  type=int, default=MAX_YT_SEARCHES_PER_RUN)
    parser.add_argument("--output",     type=str, default=OUTPUT_XLSX)
    args = parser.parse_args()
    main(args.cse_budget, args.yt_budget, args.output)
