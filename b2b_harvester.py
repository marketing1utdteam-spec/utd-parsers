#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════╗
║            IT Company Email Harvester  v4.0                        ║
║   Live-Company & Deliverable-Contact Collector                     ║
║                                                                    ║
║   v4 fixes the quality problem of v3:                              ║
║     • email-domain MUST match the company's own website domain     ║
║     • domain MUST have MX records  (checked via DNS-over-HTTPS,     ║
║       so it works even when outbound port 25 is blocked)           ║
║     • website MUST be live and NOT parked / for-sale / suspended   ║
║     • company name is taken from the HOMEPAGE, never a listicle    ║
║     • every contact gets a status (live/risky) + confidence score  ║
║                                                                    ║
║   Reusable │ Resumable │ Rotating API keys │ Excel + Google Sheets ║
╚═══════════════════════════════════════════════════════════════════╝

Usage:
    python email_harvester_v4.py                  # normal run
    python email_harvester_v4.py --reset          # clear state, start fresh
    python email_harvester_v4.py --min-score 70   # raise quality bar
    python email_harvester_v4.py --key sk-ant-..  # add Claude key (optional)
    Ctrl+C                                         # graceful stop + save

To retarget to a different niche, edit the PROFILE block below — nothing
else needs to change.
"""

import os, re, sys, json, time, signal, logging, hashlib, argparse, random
from datetime import datetime
from urllib.parse import urlparse, urljoin
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════════
#   🎯  PROFILE  —  edit THIS block to retarget the whole harvester
# ═══════════════════════════════════════════════════════════════════
#  Everything niche-specific lives here so the engine below stays generic.

PROFILE = {
    "name": "web_shopify_agencies",

    # What services the target companies sell (drives query generation).
    "services": [
        "shopify development", "shopify store setup", "shopify agency",
        "shopify theme development", "shopify plus", "shopify expert",
        "web design", "website development", "ecommerce development",
        "ecommerce agency", "online store development", "webflow agency",
        "wordpress development", "woocommerce development",
        "ui ux design", "landing page design", "conversion rate optimization",
        "headless commerce", "custom website", "digital agency",
        "shopify migration", "magento to shopify migration",
        "woocommerce to shopify", "shopify redesign",
        "shopify speed optimization", "shopify seo",
        "ecommerce web design", "dtc ecommerce agency",
        "b2b ecommerce development", "shopify app development",
        "email marketing for ecommerce", "klaviyo agency",
        "ecommerce cro", "online shop design",
        "ecommerce website redesign", "shopify maintenance",
        "shopify support agency", "subscription ecommerce development",
    ],

    # Words that, on the page, CONFIRM it is the right kind of company.
    "positive": [
        "web design", "web development", "website design", "shopify",
        "ecommerce", "e-commerce", "online store", "web agency",
        "design agency", "creative agency", "digital agency", "webflow",
        "wordpress", "woocommerce", "ui design", "ux design", "our work",
        "case studies", "portfolio", "client projects", "web studio",
        "design studio", "shopify partner", "shopify expert", "we build",
        "we design", "we develop", "conversion optimization",
        # Local-language agency terms (DE/FR/ES/IT/NL/PT/PL) so non-English
        # agencies pass classification without loosening strictness.
        "agentur", "webagentur", "webdesign", "webentwicklung", "onlineshop",
        "agence", "agencia", "agenzia", "bureau", "agência", "agencja",
    ],

    # Words that DISQUALIFY a company (wrong industry / not an agency).
    "negative": [
        "recruiting agency", "staffing agency", "talent acquisition",
        "executive search", "non-profit", "nonprofit", "charity",
        "government agency", "law firm", "legal services", "accounting firm",
        "cpa firm", "insurance company", "real estate brokerage",
        "hospital system", "medical center", "we are hiring", "job openings",
    ],

    # One-line description fed to the optional Claude validator.
    "claude_question": (
        "Does this company SELL web design, web development, ecommerce, "
        "Shopify, or digital agency services TO OTHER BUSINESSES (clients)?"
    ),
}

# ═══════════════════════════════════════════════════════════════════
#   ⚙️  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

# Claude API key for AI validation of borderline cases (optional).
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Google CSE key / id pairs — rotate on quota/block.
_API_KEYS = [k.strip() for k in os.environ.get("GOOGLE_API_KEYS", "").split(",") if k.strip()]
_CSE_IDS  = [c.strip() for c in os.environ.get("GOOGLE_CSE_IDS", "").split(",") if c.strip()]
API_PAIRS = [{"api_key": k, "cse_id": _CSE_IDS[i % len(_CSE_IDS)]}
             for i, k in enumerate(_API_KEYS)]
# KEY_SLICE ("start:end") gives this parser a DEDICATED slice of the shared key
# pool so the three parsers never compete for the same daily CSE quota (that
# contention is what caused the 429 storms in the afternoon runs). Falls back to
# the full pool if unset or misconfigured.
_KEY_SLICE = os.environ.get("KEY_SLICE", "").strip()
if _KEY_SLICE and ":" in _KEY_SLICE:
    _a, _b = _KEY_SLICE.split(":")
    API_PAIRS = API_PAIRS[(int(_a) if _a else None):(int(_b) if _b else None)] or API_PAIRS

# ─── File paths (STATE_DIR keeps logs/state on the host volume) ──
_STATE_DIR    = os.environ.get("STATE_DIR", ".")
os.makedirs(_STATE_DIR, exist_ok=True)
OUTPUT_EXCEL  = os.path.join(_STATE_DIR, "it_emails_v4.xlsx")
STATE_FILE    = os.path.join(_STATE_DIR, "harvester_v4_state.json")
VISITED_FILE  = os.path.join(_STATE_DIR, "harvester_v4_visited.json")
LOG_FILE      = os.path.join(_STATE_DIR, "harvester_v4.log")

# ─── Google Sheets ───────────────────────────────────────────────
SHEETS_SPREADSHEET_ID = os.environ.get("B2B_SHEET_ID", "")
SHEETS_WORKSHEET_GID  = int(os.environ.get("B2B_SHEET_GID", "106542427"))
SHEETS_CREDS_FILE     = os.environ.get("GOOGLE_CREDS_FILE", os.path.join(_STATE_DIR, "google_credentials.json"))

# ─── Session limits ─────────────────────────────────────────────
MAX_GOOGLE_QUERIES  = int(os.environ.get("MAX_GOOGLE_QUERIES", "50"))  # new unique Google searches per session
RESULTS_PER_QUERY   = 10    # Google results per query (max 10)
SEARCH_PAGES        = int(os.environ.get("SEARCH_PAGES", "2"))  # result pages per query (each page = 1 API call)
MAX_SUBPAGES        = 5     # extra subpages checked per domain (covers local legal pages: impressum/mentions-legales/aviso-legal)
MIN_SCORE           = 60    # keep contacts scoring >= this (0-100)

# ─── Mailbox verification (optional free-tier API) ──────────────
#  Confirms a SPECIFIC mailbox exists (SMTP + catch-all) via a provider
#  that owns port-25-enabled IPs — the only realistic way given local
#  port 25 is blocked. Leave key empty to disable (free gauntlet only).
#
#  provider = "reoon"          → https://reoon.com/  (free credits on signup)
#             "millionverifier"→ https://millionverifier.com/ (~1000 free)
#  Multiple keys are rotated round-robin to spread load and pool free
#  credits across accounts; an exhausted/invalid key auto-drops out.
#
#  Quality policy (defaults tuned for HIGHEST quality):
#    deliverable    → keep (+score)
#    undeliverable  → drop (mailbox does not exist)
#    catch-all      → drop  (can't confirm; set False to keep)
#    unknown        → drop  (couldn't verify; set False to keep)
#  When ALL credits run out the harvester does NOT stop: it falls back to
#  the free gauntlet (MX + domain + live site) and tags those rows
#  '+unverified' in the 'Validated By' column (no schema/column change).
MAILBOX_PROVIDER = os.environ.get("MAILBOX_PROVIDER", "millionverifier")  # "millionverifier" | "reoon" | "" (off)
# Keys come from env MILLIONVERIFIER_KEYS (comma-separated); fallback to the
# embedded list so it still works locally. Rotated round-robin.
MAILBOX_API_KEYS = [k.strip() for k in os.environ.get("MILLIONVERIFIER_KEYS", "").split(",") if k.strip()]
MAILBOX_DROP_CATCHALL = True   # drop catch-all domains (can't confirm mailbox)
MAILBOX_DROP_UNKNOWN  = True   # drop unverifiable results (max precision)

# ─── Delays (seconds) ───────────────────────────────────────────
DELAY_PAGE   = 1.2
DELAY_URL    = 1.6
DELAY_QUERY  = 2.4
API_ROT_WAIT = 4.0

# ─── Rotating user agents ───────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

# ═══════════════════════════════════════════════════════════════════
#   🚫  FILTERING CONSTANTS
# ═══════════════════════════════════════════════════════════════════

FREE_EMAIL_DOMAINS = {
    "gmail.com","yahoo.com","hotmail.com","outlook.com","icloud.com",
    "mail.com","protonmail.com","aol.com","yandex.ru","yandex.com",
    "zoho.com","live.com","msn.com","me.com","mac.com","inbox.com",
    "gmx.com","mail.ru","bk.ru","list.ru","inbox.ru","rambler.ru",
}

# Obvious placeholder / template / tracking addresses — hard reject.
PLACEHOLDER_RE = re.compile(
    r"your-?company|yourcompany|your-?domain|example\.|domain\.com"
    r"|email\.com|sample\.|test@|@test\.|name@|firstname|lastname"
    r"|user@|someone@|sentry|wixpress|\.png@|\.jpg@|@2x|sentry-next"
    r"|@sentry\.|@w3\.|@schema\.|@wordpress\.|@shopify\.com|@apple\.com"
    r"|@google\.com|@amazon\.com|@microsoft\.com|@facebook\.com"
    r"|sarah@startup\.com|john@|jane@|hello@example",
    re.IGNORECASE,
)

BAD_EMAIL_RE = re.compile(
    r"noreply|no-reply|donotreply|do-not-reply|unsubscribe"
    r"|privacy@|legal@|press@|media@|abuse@|spam@|postmaster@"
    r"|hostmaster@|mailer-daemon@|webmaster@",
    re.IGNORECASE,
)

LARGE_CORP_DOMAINS = {
    "accenture.com","deloitte.com","ibm.com","cognizant.com","infosys.com",
    "wipro.com","tcs.com","capgemini.com","hcltech.com","techmahindra.com",
    "salesforce.com","oracle.com","sap.com","adobe.com","hubspot.com",
    "shopify.com","bigcommerce.com","magento.com","squarespace.com",
    "wix.com","godaddy.com","wordpress.com","kpmg.com","pwc.com","ey.com",
}

DIRECTORY_DOMAINS = {
    "clutch.co","goodfirms.co","designrush.com","sortlist.com","manifest.co",
    "expertise.com","upcity.com","g2.com","capterra.com","trustpilot.com",
    "yelp.com","awwwards.com","cssdesignawards.com","webbyawards.com",
    "linkedin.com","twitter.com","x.com","facebook.com","instagram.com",
    "youtube.com","tiktok.com","pinterest.com","reddit.com","quora.com",
    "techcrunch.com","wired.com","forbes.com","medium.com","substack.com",
    "upwork.com","fiverr.com","toptal.com","freelancer.com","guru.com",
    "wadline.com","themanifest.com","github.com","wikipedia.org",
}

GOV_EDU_TLDS = {
    ".gov",".mil",".edu",".ac.uk",".gov.uk",".gov.au",
    ".gov.ca",".gc.ca",".gouv.fr",".gob.es",".gov.ie",
}

# Page-path fragments that mean "this is an article/listicle, not a company".
JUNK_PATH = [
    "/blog/", "/news/", "/article", "/sitemap", "/html-sitemap",
    "/tag/", "/category/", "/author/", "/wp-json", "/feed",
]
JUNK_TITLE = [
    "sitemap", "top 10", "top 20", "best ", " vs ", "how to",
    "ultimate guide", "listicle", "privacy policy", "terms of",
    "404", "not found", "page not found", "blog -", "- blog",
]

# Signals that a domain is dead / parked / for-sale / suspended.
DEAD_SIGNALS = [
    "domain is for sale", "buy this domain", "this domain is for sale",
    "domain for sale", "is for sale", "parked free", "sedoparking",
    "this domain is parked", "domain parking", "hugedomains",
    "account suspended", "this account has been suspended",
    "site temporarily unavailable", "website is temporarily unavailable",
    "default web page", "welcome to nginx", "apache2 ubuntu default",
    "future home of something", "index of /",
]

# ─── Email priority (role quality) ───────────────────────────────
EMAIL_PRIORITY = [
    "contact","hello","hi","hey","info","team","studio","agency",
    "work","hire","projects","enquiry","enquiries","newbusiness",
    "business","office","grow","connect","bonjour",
]
EMAIL_DEPRIO = {
    "support","help","sales","billing","invoice","jobs","career",
    "careers","hr","recruiting","newsletter","subscribe","admin",
    "accounts","accounting","finance","noreply",
}

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]{1,64}@[a-zA-Z0-9.\-]{1,253}\.[a-zA-Z]{2,12}",
    re.IGNORECASE,
)

# Generic multi-part TLDs needed to compute the registrable domain.
MULTI_TLDS = {
    "co.uk","org.uk","gov.uk","ac.uk","com.au","net.au","org.au",
    "co.nz","com.br","co.za","co.in","com.sg","com.hk","co.jp",
    "com.mx","com.tr","co.il",
}

# ═══════════════════════════════════════════════════════════════════
#   📋  LOGGING
# ═══════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════
#   🌐  DOMAIN HELPERS
# ═══════════════════════════════════════════════════════════════════

def registrable_domain(host_or_url: str) -> str:
    """example.com from www.sub.example.com; handles co.uk-style TLDs."""
    h = host_or_url.strip().lower()
    if "://" in h:
        h = urlparse(h).netloc
    h = h.split("/")[0].split("@")[-1].replace("www.", "")
    if not h or "." not in h:
        return h
    parts = h.split(".")
    last2 = ".".join(parts[-2:])
    last3 = ".".join(parts[-3:]) if len(parts) >= 3 else ""
    if last2 in MULTI_TLDS and len(parts) >= 3:
        return ".".join(parts[-3:])
    return last2


def root_url(url: str) -> str:
    """Normalise any deep URL to its homepage (scheme://host)."""
    try:
        p = urlparse(url)
        if not p.scheme or not p.netloc:
            return ""
        return f"{p.scheme}://{p.netloc}"
    except Exception:
        return ""

# ═══════════════════════════════════════════════════════════════════
#   📨  MX RESOLVER  (DNS-over-HTTPS — works without port 25)
# ═══════════════════════════════════════════════════════════════════

class MXResolver:
    """
    Confirms a domain can receive email by checking for MX (or A as a
    fallback) records via DNS-over-HTTPS. Uses Cloudflare then Google,
    so it works even on networks that block outbound SMTP (port 25).
    Results are cached per domain for the whole run.
    """

    _DOH = [
        ("https://cloudflare-dns.com/dns-query", {"accept": "application/dns-json"}),
        ("https://dns.google/resolve", {}),
    ]

    def __init__(self):
        self.cache = {}     # domain -> bool
        self.http  = requests.Session()

    def _query(self, domain: str, rtype: str) -> list:
        for base, hdrs in self._DOH:
            try:
                r = self.http.get(
                    base, params={"name": domain, "type": rtype},
                    headers=hdrs, timeout=10,
                )
                if r.status_code == 200:
                    return r.json().get("Answer", []) or []
            except Exception:
                continue
        return None  # network failure (distinct from "no records")

    def has_mail(self, domain: str) -> bool:
        domain = registrable_domain(domain)
        if not domain:
            return False
        if domain in self.cache:
            return self.cache[domain]

        mx = self._query(domain, "MX")
        if mx is None:                       # DoH failed → don't punish domain
            return True
        ok = len(mx) > 0
        if not ok:                           # no MX: many small sites use A for mail
            a = self._query(domain, "A")
            ok = bool(a)
        self.cache[domain] = ok
        return ok

# ═══════════════════════════════════════════════════════════════════
#   📧  EMAIL UTILITIES
# ═══════════════════════════════════════════════════════════════════

def is_valid_email(em: str) -> bool:
    em = em.lower().strip()
    if not em or "@" not in em or len(em) > 254:
        return False
    local, domain = em.rsplit("@", 1)
    if not local or not domain or "." not in domain:
        return False
    if domain in FREE_EMAIL_DOMAINS:
        return False
    if PLACEHOLDER_RE.search(em) or BAD_EMAIL_RE.search(em):
        return False
    # reject image/asset filenames that look like emails
    if re.search(r"\.(png|jpg|jpeg|gif|svg|webp|css|js)$", domain):
        return False
    tld_len = len(domain.split(".")[-1])
    return 2 <= tld_len <= 12


def extract_emails(text: str) -> set:
    return {e.lower() for e in EMAIL_RE.findall(text) if is_valid_email(e)}


def email_role_score(local: str) -> int:
    base = local.split("@")[0].replace(".", "").replace("-", "").replace("_", "")
    first = local.split("@")[0].split(".")[0]
    if base in EMAIL_PRIORITY:
        return 25
    if first in EMAIL_DEPRIO:
        return 6
    return 16  # personal name e.g. john@ — fine for outreach


def pick_best_email(emails: set, site_domain: str) -> str:
    """Prefer on-domain, role-priority addresses."""
    on_domain = [e for e in emails
                 if registrable_domain(e.split("@")[-1]) == site_domain]
    pool = on_domain or list(emails)
    if not pool:
        return ""
    return sorted(pool, key=lambda e: (-email_role_score(e.split("@")[0]), e))[0]

# ═══════════════════════════════════════════════════════════════════
#   🏢  URL PRE-FILTER
# ═══════════════════════════════════════════════════════════════════

def is_company_url(url: str) -> bool:
    """Reject directories, big corps, gov/edu, and obvious article paths."""
    try:
        p = urlparse(url)
        domain = p.netloc.lower().replace("www.", "")
        path   = p.path.lower()
    except Exception:
        return False
    if not domain:
        return False
    if domain in LARGE_CORP_DOMAINS or domain in DIRECTORY_DOMAINS:
        return False
    if any(domain.endswith(t) for t in GOV_EDU_TLDS):
        return False
    if any(j in path for j in JUNK_PATH):
        return False
    return True

# ═══════════════════════════════════════════════════════════════════
#   ✅  VALIDATORS  (is it the right KIND of company?)
# ═══════════════════════════════════════════════════════════════════

class RuleValidator:
    def classify(self, company: str, url: str, page_text: str) -> str:
        blob = (company + " " + url + " " + page_text[:4000]).lower()
        neg  = sum(1 for kw in PROFILE["negative"] if kw in blob)
        pos  = sum(1 for kw in PROFILE["positive"] if kw in blob)
        if neg >= 2:              return "reject"
        if neg >= 1 and pos == 0: return "reject"
        if pos >= 3:              return "accept"
        if pos >= 1 and neg == 0: return "accept"
        return "uncertain"


class ClaudeValidator:
    _API = "https://api.anthropic.com/v1/messages"
    _MDL = "claude-haiku-4-5-20251001"

    def __init__(self, key: str):
        self.key   = (key or "").strip()
        self._cache = {}

    def enabled(self) -> bool:
        return bool(self.key)

    def classify(self, company: str, url: str, snippet: str) -> str:
        if not self.key:
            return "unknown"
        dom = registrable_domain(url)
        if dom in self._cache:
            return self._cache[dom]
        try:
            r = self._call(company, url, snippet[:600])
            self._cache[dom] = r
            return r
        except Exception as e:
            log.debug(f"Claude error: {e}")
            return "unknown"

    def _call(self, company: str, url: str, snippet: str) -> str:
        hdrs = {"x-api-key": self.key, "anthropic-version": "2023-06-01",
                "content-type": "application/json"}
        prompt = (
            f"{PROFILE['claude_question']}\n\n"
            f"Company: {company}\nURL: {url}\nPage excerpt: {snippet}\n\n"
            "Reply with exactly one word:\n"
            "ACCEPT — yes, it matches.\n"
            "REJECT — no (wrong industry, recruiter, nonprofit, government, "
            "SaaS product, marketplace, or unrelated)."
        )
        body = {"model": self._MDL, "max_tokens": 10,
                "messages": [{"role": "user", "content": prompt}]}
        resp = requests.post(self._API, headers=hdrs, json=body, timeout=22)
        if resp.status_code == 200:
            txt = resp.json()["content"][0]["text"].strip().upper()
            if "ACCEPT" in txt: return "accept"
            if "REJECT" in txt: return "reject"
        elif resp.status_code == 529:
            time.sleep(8)
        return "unknown"


class CompanyValidator:
    def __init__(self, anthropic_key: str = ""):
        self.rules  = RuleValidator()
        self.claude = ClaudeValidator(anthropic_key)
        self.stats  = {"accepted": 0, "rejected": 0,
                       "claude_calls": 0, "uncertain_dropped": 0}

    def validate(self, company: str, url: str, page_text: str) -> tuple:
        result = self.rules.classify(company, url, page_text)
        if result == "reject":
            self.stats["rejected"] += 1
            return False, "rules"
        if result == "accept":
            self.stats["accepted"] += 1
            return True, "rules"
        if self.claude.enabled():
            self.stats["claude_calls"] += 1
            cr = self.claude.classify(company, url, page_text)
            if cr == "accept":
                self.stats["accepted"] += 1
                return True, "claude"
            if cr == "reject":
                self.stats["rejected"] += 1
                return False, "claude"
        self.stats["uncertain_dropped"] += 1
        return False, "uncertain"

# ═══════════════════════════════════════════════════════════════════
#   🔎  CONTACT VERIFIER  (is the company alive & email deliverable?)
# ═══════════════════════════════════════════════════════════════════

class ContactVerifier:
    """
    The heart of v4. Produces (status, score, reason) for an email.
      status: 'live'  (keep, high confidence)
              'risky' (keep only if below MIN_SCORE not enforced)
              'dead'  (reject)
    Free signals only — no port-25 SMTP needed.
    """

    def __init__(self, mx: MXResolver):
        self.mx = mx
        self.stats = {"no_mx": 0, "mismatch": 0, "bad_format": 0,
                      "dead_site": 0, "live": 0, "risky": 0}

    def verify(self, email: str, site_domain: str,
               page_text: str, site_alive: bool) -> tuple:
        email = email.lower().strip()

        # 1. Format / placeholder
        if not is_valid_email(email):
            self.stats["bad_format"] += 1
            return "dead", 0, "bad_format"

        edom = registrable_domain(email.split("@")[-1])

        # 2. Email must live on the company's OWN domain (kills 3rd-party junk)
        if site_domain and edom != site_domain:
            self.stats["mismatch"] += 1
            return "dead", 0, "domain_mismatch"

        # 3. Domain must be able to receive mail (MX via DoH)
        if not self.mx.has_mail(edom):
            self.stats["no_mx"] += 1
            return "dead", 0, "no_mx"

        # 4. Website must be alive (checked upstream during scrape)
        if not site_alive:
            self.stats["dead_site"] += 1
            return "dead", 0, "dead_site"

        # ── Scoring (0-100) ──────────────────────────────────────
        score = 40                                   # MX present + on-domain
        score += 20                                  # site live
        score += email_role_score(email.split("@")[0])   # role quality 6-25

        # freshness bonus: recent year somewhere on the page
        try:
            yrs = [int(y) for y in re.findall(r"\b(20[12]\d)\b", page_text or "")]
            this_year = datetime.now().year
            if yrs and max(yrs) >= this_year - 1:
                score += 15
        except Exception:
            pass

        score = min(score, 100)
        status = "live" if score >= MIN_SCORE else "risky"
        self.stats[status] += 1
        return status, score, "ok"

# ═══════════════════════════════════════════════════════════════════
#   ✉️  MAILBOX VERIFIER  (optional free-tier API — real RCPT check)
# ═══════════════════════════════════════════════════════════════════

class MailboxVerifier:
    """
    Confirms a specific mailbox exists via a hosted verification API
    (MillionVerifier or Reoon). These services own port-25-enabled IPs
    with sending reputation, so they return reliable SMTP/catch-all
    results that a local probe (port 25 blocked) cannot.

    Multiple API keys are rotated ROUND-ROBIN so load and free credits
    are spread evenly across accounts. A key that returns an
    exhausted/invalid error is dropped from rotation for the rest of
    the run; the call then retries with the next live key.

    Canonical statuses returned:
        'deliverable' | 'catch_all' | 'undeliverable' | 'unknown' | 'off'
    Caches per email; fails safe to 'unknown' on any error.
    """

    def __init__(self, provider: str, keys):
        self.provider = (provider or "").strip().lower()
        if isinstance(keys, str):
            keys = [keys]
        self.keys = [k.strip() for k in (keys or []) if k and k.strip()]
        self.idx  = 0
        self.dead = set()              # indices of exhausted/invalid keys
        self.http = requests.Session()
        self.cache = {}
        self.per_key = [0] * len(self.keys)   # calls per key (even-load proof)
        self.stats = {"deliverable": 0, "catch_all": 0,
                      "undeliverable": 0, "unknown": 0, "calls": 0}

    def enabled(self) -> bool:
        return bool(self.provider and self.keys)

    def exhausted(self) -> bool:
        """True once every key has been dropped (out of credits/invalid)."""
        return self.enabled() and len(self.dead) >= len(self.keys)

    def _next_key_idx(self):
        """Round-robin index over keys that are still alive."""
        for _ in range(len(self.keys)):
            i = self.idx % len(self.keys)
            self.idx += 1
            if i not in self.dead:
                return i
        return None

    def verify(self, email: str) -> str:
        if not self.enabled():
            return "off"
        email = email.lower().strip()
        if email in self.cache:
            return self.cache[email]

        res = "unknown"
        for _ in range(len(self.keys)):
            i = self._next_key_idx()
            if i is None:
                log.warning("  ✉️  all mailbox keys exhausted — skipping check")
                break
            key = self.keys[i]
            self.per_key[i] += 1
            try:
                if self.provider == "reoon":
                    status, failed, exhaust = self._reoon(email, key)
                elif self.provider == "millionverifier":
                    status, failed, exhaust = self._millionverifier(email, key)
                else:
                    log.warning(f"  unknown mailbox provider '{self.provider}'")
                    break
            except Exception as e:
                log.debug(f"  mailbox API error: {e}")
                status, failed, exhaust = "unknown", True, False

            if exhaust:
                self.dead.add(i)
                log.warning(f"  🔄 mailbox key #{i+1} (…{key[-4:]}) "
                            f"exhausted/invalid — dropped from rotation")
            if failed:
                continue          # try next key
            res = status
            break

        self.stats["calls"] += 1
        self.stats[res] = self.stats.get(res, 0) + 1
        self.cache[email] = res
        return res

    # Each provider returns (status, failed, exhaust):
    #   failed  → this call produced no usable result, rotate & retry
    #   exhaust → this key is out of credits / invalid, drop it for good

    # ── MillionVerifier: api.millionverifier.com/api/v3 ─────────
    def _millionverifier(self, email: str, key: str):
        r = self.http.get(
            "https://api.millionverifier.com/api/v3/",
            params={"api": key, "email": email, "timeout": 20},
            timeout=35,
        )
        if r.status_code in (401, 403):
            return "unknown", True, True
        if r.status_code == 429:
            return "unknown", True, False     # rate-limited, key still good
        if r.status_code != 200:
            return "unknown", True, False
        d = r.json()
        err = str(d.get("error", "")).lower()
        if err:
            kill = any(w in err for w in ("credit", "api key", "invalid", "key"))
            return "unknown", True, kill
        res = str(d.get("result", "")).lower()
        credits = d.get("credits")
        out_of_credits = isinstance(credits, (int, float)) and credits <= 0
        if res == "ok":
            return "deliverable", False, out_of_credits
        if res in ("catch_all", "catch-all"):
            return "catch_all", False, out_of_credits
        if res in ("invalid", "disposable", "error"):
            return "undeliverable", False, out_of_credits
        return "unknown", False, out_of_credits

    # ── Reoon: emailverifier.reoon.com/api/v1 ───────────────────
    def _reoon(self, email: str, key: str):
        r = self.http.get(
            "https://emailverifier.reoon.com/api/v1/verify",
            params={"email": email, "key": key, "mode": "power"},
            timeout=35,
        )
        if r.status_code in (401, 403):
            return "unknown", True, True
        if r.status_code != 200:
            return "unknown", True, False
        d = r.json()
        st = str(d.get("status", "")).lower()
        if d.get("is_safe_to_send") is True or st in ("safe", "valid"):
            return "deliverable", False, False
        if st in ("invalid", "disabled", "disposable", "spamtrap", "undeliverable"):
            return "undeliverable", False, False
        if st in ("catch_all", "catch-all", "accept_all"):
            return "catch_all", False, False
        return "unknown", False, False


# ═══════════════════════════════════════════════════════════════════
#   🔄  DYNAMIC QUERY GENERATOR
# ═══════════════════════════════════════════════════════════════════

class QueryGenerator:
    """Unlimited unique, niche-targeted Google queries (deterministic).

    v4.1 — MULTILINGUAL. English tight queries have dried up (Google returns
    ~2 results for tight EN queries), while non-English markets are a huge
    fresh vein (DE "shopify agentur" impressum ≈15k, ES "agencia shopify"
    contacto ≈9k, DE webagentur ≈5.5k — all REAL agencies). Each generated
    query now carries a language tag (lang_de/fr/es/... or lang_en) that the
    GoogleClient turns into the `lr` restrict-language parameter. Local
    legal/contact pages (Impressum, mentions légales, aviso legal) are
    required to carry an email by law → strongest possible signal.

    Every generated query is still md5-hashed and deduped via used_hashes so
    the deterministic counter mechanism is preserved unchanged.
    """

    _TYPE = ["agency", "studio", "company", "team", "design studio",
             "development studio", "digital agency", "web studio",
             "creative agency", "web agency", "consultancy", "experts",
             "developers", "partners"]

    _LOC = [
        "", "", "", "", "", "",
        # North America
        "USA", "New York", "Los Angeles", "Chicago", "San Francisco",
        "Austin", "Seattle", "Boston", "Miami", "Denver", "Dallas",
        "Atlanta", "Portland", "San Diego", "Phoenix", "Houston",
        "Philadelphia", "Minneapolis", "Nashville", "Charlotte",
        "Orlando", "Tampa", "Las Vegas", "Columbus", "Kansas City",
        "Indianapolis", "Detroit", "Pittsburgh", "Salt Lake City",
        "Canada", "Toronto", "Vancouver", "Montreal", "Calgary", "Ottawa",
        # UK & Ireland
        "UK", "London", "Manchester", "Birmingham", "Bristol",
        "Leeds", "Glasgow", "Edinburgh", "Liverpool", "Nottingham",
        "Ireland", "Dublin",
        # Australia / NZ / Singapore (English-speaking APAC)
        "Australia", "Sydney", "Melbourne", "Brisbane", "Perth", "Adelaide",
        "Singapore", "New Zealand", "Auckland",
    ]

    # contact-intent suffix + EXCLUDE article/marketplace noise (ENGLISH).
    _EXCLUDE = '-"top 10" -"best" -blog -listicle -clutch -upwork -fiverr -jobs -careers'
    _SUFFIX = [
        f'"contact us" ("info@" OR "hello@") {_EXCLUDE}',
        f'contact email {_EXCLUDE}',
        f'"get in touch" email {_EXCLUDE}',
        f'"work with us" email {_EXCLUDE}',
        f'("hello@" OR "contact@") {_EXCLUDE}',
        f'"request a quote" ("info@" OR "hello@") {_EXCLUDE}',
        f'"start a project" email {_EXCLUDE}',
        f'"free consultation" ("hello@" OR "contact@") {_EXCLUDE}',
        f'"book a call" ("hello@" OR "info@") {_EXCLUDE}',
        f'"shopify partner" contact {_EXCLUDE}',
    ]

    # ── LOCAL-LANGUAGE query pools ──────────────────────────────────────
    # Each entry is a ready-made tight query template. {c} is filled with a
    # local city (or dropped). These are the fresh, deep veins. Local
    # contact/legal words (Impressum, Kontakt, mentions légales, aviso legal,
    # contatti, contacto...) are legally required to carry an email address.
    _LOCAL = {
        "lang_de": {
            "cities": ["Berlin", "München", "Hamburg", "Köln", "Frankfurt",
                       "Stuttgart", "Düsseldorf", "Wien", "Zürich", "Leipzig"],
            "templates": [
                '"shopify agentur" impressum',
                '"shopify agentur" kontakt',
                '"webagentur" "shopify" impressum',
                '"shopify agentur" {c} impressum',
                '"shopify agentur" {c} kontakt',
                '"e-commerce agentur" shopify impressum',
                '"webagentur" {c} shopify kontakt',
                '"shopify partner" agentur impressum',
                '"onlineshop agentur" shopify kontakt',
            ],
        },
        "lang_fr": {
            "cities": ["Paris", "Lyon", "Marseille", "Bordeaux", "Toulouse",
                       "Lille", "Nantes", "Bruxelles", "Genève"],
            "templates": [
                '"agence shopify" contact',
                '"agence shopify" "mentions légales"',
                '"agence e-commerce shopify" contact',
                '"agence shopify" {c} contact',
                '"agence shopify" {c} "mentions légales"',
                '"agence web" shopify "mentions légales"',
                '"shopify partner" agence contact',
            ],
        },
        "lang_es": {
            "cities": ["Madrid", "Barcelona", "Valencia", "Sevilla",
                       "Málaga", "Bilbao", "Zaragoza"],
            "templates": [
                '"agencia shopify" contacto',
                '"agencia shopify" "aviso legal"',
                '"agencia ecommerce shopify" contacto',
                '"agencia shopify" {c} contacto',
                '"agencia shopify" {c}',
                '"agencia ecommerce" shopify "aviso legal"',
                '"shopify partner" agencia contacto',
            ],
        },
        "lang_it": {
            "cities": ["Milano", "Roma", "Torino", "Napoli", "Bologna",
                       "Firenze", "Verona"],
            "templates": [
                '"agenzia shopify" contatti',
                '"agenzia shopify" contattaci',
                '"agenzia shopify" {c} contatti',
                '"agenzia e-commerce shopify" contatti',
                '"shopify partner" agenzia contatti',
            ],
        },
        "lang_nl": {
            "cities": ["Amsterdam", "Rotterdam", "Utrecht", "Den Haag",
                       "Eindhoven", "Antwerpen", "Gent"],
            "templates": [
                '"shopify bureau" contact',
                '"shopify webshop laten maken"',
                '"shopify bureau" {c} contact',
                '"e-commerce bureau" shopify contact',
                '"shopify partner" bureau contact',
            ],
        },
        "lang_pt": {
            "cities": ["Lisboa", "Porto", "Braga", "São Paulo",
                       "Rio de Janeiro", "Lisbon"],
            "templates": [
                '"agência shopify" contato',
                '"agência shopify" contacto',
                '"agência shopify" {c} contato',
                '"agência e-commerce" shopify contato',
                '"shopify partner" agência contato',
            ],
        },
        "lang_pl": {
            "cities": ["Warszawa", "Kraków", "Wrocław", "Poznań",
                       "Gdańsk", "Łódź"],
            "templates": [
                '"agencja shopify" kontakt',
                '"agencja shopify" {c} kontakt',
                '"agencja e-commerce" shopify kontakt',
                '"shopify partner" agencja kontakt',
            ],
        },
    }
    # Languages weighted so local (fresh) markets dominate; English stays in
    # the mix but is a minority. Repetition = higher probability.
    _LANG_POOL = (
        ["lang_de"] * 5 + ["lang_fr"] * 4 + ["lang_es"] * 4 +
        ["lang_it"] * 3 + ["lang_nl"] * 3 + ["lang_pt"] * 2 +
        ["lang_pl"] * 2 + ["lang_en"] * 4
    )

    def __init__(self, counter: int, used_hashes: set):
        self.counter     = counter
        self.used_hashes = used_hashes

    def _hash(self, q: str) -> str:
        return hashlib.md5(q.lower().strip().encode()).hexdigest()[:14]

    def _build_en(self, rng) -> str:
        svc = rng.choice(PROFILE["services"])
        typ = rng.choice(self._TYPE)
        loc = rng.choice(self._LOC)
        sfx = rng.choice(self._SUFFIX)
        loc_s = f"{loc} " if loc else ""
        v = rng.randint(0, 3)
        if v == 0:   q = f'"{svc} {typ}" {loc_s}{sfx}'
        elif v == 1: q = f'"{svc}" {typ} {loc_s}{sfx}'
        elif v == 2: q = f'"{svc} {typ}" {sfx}'
        else:        q = f'"{svc}" agency {loc_s}{sfx}'
        return re.sub(r"\s{2,}", " ", q).strip()

    def _build_local(self, rng, lang: str) -> str:
        pool = self._LOCAL[lang]
        tpl  = rng.choice(pool["templates"])
        if "{c}" in tpl:
            tpl = tpl.replace("{c}", rng.choice(pool["cities"]))
        return re.sub(r"\s{2,}", " ", tpl).strip()

    def _build(self) -> tuple:
        """Return (query, lang). lang is an `lr` code like 'lang_de'."""
        rng = random.Random(self.counter)
        self.counter += 1
        lang = rng.choice(self._LANG_POOL)
        if lang == "lang_en":
            return self._build_en(rng), "lang_en"
        return self._build_local(rng, lang), lang

    def next_batch(self, n: int) -> list:
        """Return a list of (query, lang) tuples, deduped by query hash."""
        queries, attempts = [], 0
        while len(queries) < n and attempts < n * 300:
            q, lang = self._build()
            h = self._hash(q)
            if h not in self.used_hashes and len(q) > 8:
                self.used_hashes.add(h)
                queries.append((q, lang))
            attempts += 1
        return queries

# ═══════════════════════════════════════════════════════════════════
#   🔑  GOOGLE SEARCH CLIENT  (rotating keys)
# ═══════════════════════════════════════════════════════════════════

class GoogleClient:
    _BASE = "https://www.googleapis.com/customsearch/v1"

    def __init__(self, pairs: list):
        self.pairs    = pairs
        self.idx      = 0
        self.failures = [0] * len(pairs)
        self.dead     = set()          # pairs out of daily quota (429/403)
        self.calls    = 0

    def exhausted(self) -> bool:
        """True once every key pair is out of quota for the day. The harvester
        checks this and stops early instead of storming thousands of 429s."""
        return len(self.dead) >= len(self.pairs)

    def _drop(self, reason: str):
        """Quota/forbidden → this key is done for the run; drop it for good."""
        self.dead.add(self.idx)
        log.warning(f"  🔻 API pair #{self.idx} dropped ({reason}); "
                    f"{len(self.pairs) - len(self.dead)} live")
        self._advance()

    def _advance(self):
        for _ in range(len(self.pairs)):
            self.idx = (self.idx + 1) % len(self.pairs)
            if self.idx not in self.dead:
                return

    def _rotate(self, reason: str):
        self.failures[self.idx] += 1
        self._advance()
        time.sleep(API_ROT_WAIT)

    @staticmethod
    def _is_daily_limit(resp) -> bool:
        """Distinguish a DAILY quota kill (drop the key for the day) from a
        transient per-minute rate limit (just pause and try another key — the
        key is NOT out of quota). Marking a rate-limited key dead would falsely
        burn healthy keys during a burst."""
        try:
            body = resp.text.lower()
        except Exception:
            return False
        return any(k in body for k in (
            "dailylimitexceeded", "quotaexceeded", "quota exceeded", "daily limit"))

    def search(self, query: str, start: int = 1, lang: str = None) -> list:
        # `lang` is an `lr` restrict-language code (e.g. 'lang_de'). When set,
        # results are restricted to that language — essential for the fresh
        # non-English veins (DE/FR/ES/...). When None, no lr → widest results.
        if self.exhausted():
            return []
        for _ in range(len(self.pairs)):
            if self.exhausted():
                break
            if self.idx in self.dead:
                self._advance(); continue
            pair = self.pairs[self.idx]
            try:
                params = {"key": pair["api_key"], "cx": pair["cse_id"],
                          "q": query, "num": RESULTS_PER_QUERY,
                          "start": start}
                if lang:
                    params["lr"] = lang
                    params["hl"] = lang.replace("lang_", "")  # UI hint hint
                resp = requests.get(
                    self._BASE,
                    params=params,
                    timeout=15,
                )
                self.calls += 1
                if resp.status_code == 200:
                    return [it["link"] for it in resp.json().get("items", []) if "link" in it]
                if resp.status_code in (429, 403):
                    if self._is_daily_limit(resp):
                        self._drop(f"HTTP {resp.status_code} daily")
                    else:
                        # Transient rate limit — pause and try another key,
                        # but keep this one alive (it isn't out of daily quota).
                        self._rotate(f"HTTP {resp.status_code} rate")
                elif resp.status_code == 400:
                    return []
                else:
                    self._rotate(f"HTTP {resp.status_code}")
            except requests.Timeout:
                self._rotate("timeout")
            except requests.RequestException as e:
                self._rotate(str(e)[:40])
        if self.exhausted():
            log.error("  ❌ All API pairs out of quota — stopping search.")
        return []

# ═══════════════════════════════════════════════════════════════════
#   🌐  COMPANY WEBSITE SCRAPER  (homepage-anchored)
# ═══════════════════════════════════════════════════════════════════

def _headers() -> dict:
    return {
        "User-Agent":      random.choice(USER_AGENTS),
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate",
        "Connection":      "keep-alive",
    }

# Ordered by extraction value. Local legal/contact pages are interleaved
# high because, by law, they must carry a real email address — the strongest
# extraction targets on non-EN sites. Only the first MAX_SUBPAGES are fetched
# per domain, so highest-signal paths come first.
_SUBPAGES = [
    "/contact", "/impressum", "/kontakt", "/contact-us",
    "/mentions-legales", "/aviso-legal", "/contacto", "/contatti",
    "/contato", "/get-in-touch", "/imprint", "/about", "/about-us",
]


def _company_name(soup: BeautifulSoup, url: str) -> str:
    og = soup.find("meta", property="og:site_name")
    if og and og.get("content", "").strip():
        return og["content"].strip()[:80]
    t = soup.find("title")
    if t:
        txt = t.get_text(strip=True)
        for sep in [" | ", " – ", " — ", " - ", " · ", " :: "]:
            if sep in txt:
                return txt.split(sep)[0].strip()[:80]
        return txt[:80]
    return registrable_domain(url)


def _looks_like_junk(name: str) -> bool:
    n = name.lower()
    return any(k in n for k in JUNK_TITLE)


def scrape_company(home_url: str, session: requests.Session,
                   visited: set) -> dict:
    """
    Scrape a company HOMEPAGE (+ a few contact subpages) for one best email.
    Returns dict: {email, company, domain, page_text, alive, url}.
    'alive' is False if the site is down/parked/suspended.
    """
    out = {"email": "", "company": "", "domain": "",
           "page_text": "", "alive": False, "url": home_url}

    root = root_url(home_url)
    if not root:
        return out
    domain = registrable_domain(root)
    out["domain"] = domain

    pages = [root] + [root + sp for sp in _SUBPAGES]
    mailto_emails, text_emails = set(), set()
    tried = 0

    for url in pages:
        if not url:
            continue
        uh = hash_key(url)                # visited stores URL hashes (repo is public)
        if uh in visited:
            continue
        if tried >= MAX_SUBPAGES + 1:
            break
        visited.add(uh)
        tried += 1
        try:
            resp = session.get(url, headers=_headers(), timeout=13,
                               allow_redirects=True)
            if resp.status_code != 200:
                continue
            if "text/html" not in resp.headers.get("Content-Type", ""):
                continue

            soup = BeautifulSoup(resp.text, "html.parser")
            body = soup.get_text(" ", strip=True)

            if tried == 1:  # homepage: liveness + identity
                bl = body.lower()
                if len(bl) < 200 or any(s in bl for s in DEAD_SIGNALS):
                    log.info(f"    ⛔ dead/parked: {domain}")
                    return out                       # alive stays False
                out["alive"]     = True
                out["company"]   = _company_name(soup, url)
                out["page_text"] = body[:3000]

            for a in soup.find_all("a", href=True):
                h = a["href"]
                if h.lower().startswith("mailto:"):
                    em = h[7:].split("?")[0].strip().lower()
                    if is_valid_email(em):
                        mailto_emails.add(em)
            text_emails.update(extract_emails(resp.text))

            best = pick_best_email(mailto_emails, domain)
            if best and email_role_score(best.split("@")[0]) == 25:
                break
            if tried > 1:
                time.sleep(DELAY_PAGE)
        except requests.Timeout:
            log.debug(f"  timeout: {url}")
        except Exception as e:
            log.debug(f"  fetch error {url}: {e}")

    out["email"] = (pick_best_email(mailto_emails, domain)
                    or pick_best_email(text_emails, domain))
    return out

# ═══════════════════════════════════════════════════════════════════
#   📊  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════

_DARK, _EVEN, _ODD = "1B3A6B", "EEF4FB", "FFFFFF"
COLS   = ["Email", "Company", "Website", "Status", "Date Added"]
WIDTHS = [38, 32, 46, 22, 18]


def _border():
    s = Side(style="thin", color="C5D5E8")
    return Border(left=s, right=s, top=s, bottom=s)


def init_excel(path: str):
    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path); wb.close(); return
        except Exception:
            os.remove(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live Agency Contacts"
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color=_DARK, end_color=_DARK, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 26
    wb.save(path)
    log.info(f"📗 Excel created: {path}")


def _status_cell(r: dict) -> str:
    return f"{r['status']}·{r['score']}·{r['method']}"


def append_to_excel(path: str, records: list):
    if not records:
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    start = ws.max_row + 1
    for i, r in enumerate(records):
        row = start + i
        fc = _EVEN if row % 2 == 0 else _ODD
        fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
        vals = [r["email"], r["company"], r["url"], _status_cell(r), r["date"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(vertical="center")
            c.border = _border()
        ws.row_dimensions[row].height = 18
    wb.save(path)
    log.info(f"  📗 +{len(records)} rows  (file: {path})")

# ═══════════════════════════════════════════════════════════════════
#   🟢  GOOGLE SHEETS UPLOADER  (5-column, drop-in compatible)
# ═══════════════════════════════════════════════════════════════════

class SheetsUploader:
    def __init__(self, spreadsheet_id, gid, creds_file):
        self.spreadsheet_id = spreadsheet_id
        self.gid            = gid
        self.creds_file     = creds_file
        self._ws = None; self._ok = False; self._sent = 0

    def connect(self) -> bool:
        if not os.path.exists(self.creds_file):
            log.warning(f"⚠️  Sheets creds not found ({self.creds_file}); Excel only.")
            return False
        try:
            import gspread
        except ImportError:
            log.error("⚠️  gspread not installed; Excel only. "
                      "pip install gspread google-auth google-auth-oauthlib")
            return False
        try:
            with open(self.creds_file, "r", encoding="utf-8") as f:
                meta = json.load(f)
            if meta.get("type") == "service_account":
                gc = gspread.service_account(filename=self.creds_file)
            elif "installed" in meta or "web" in meta:
                gc = gspread.oauth(credentials_filename=self.creds_file,
                                   authorized_user_filename="google_token.json")
            else:
                log.error("⚠️  Unrecognised creds format; Excel only."); return False
            ss = gc.open_by_key(self.spreadsheet_id)
            self._ws = ss.get_worksheet_by_id(self.gid) or ss.get_worksheet(0)
            if self._ws is None:
                log.error("  No worksheet found."); return False
            self._ok = True
            log.info(f"  ✅ Google Sheets ready: tab='{self._ws.title}' in '{ss.title}'")
            return True
        except Exception as e:
            log.error(f"⚠️  Sheets connect error: {e}; Excel only.")
            return False

    def append_rows(self, records: list) -> bool:
        if not self._ok or not self._ws or not records:
            return False
        rows = [[r["email"], r["company"], r["url"],
                 _status_cell(r), r["date"]] for r in records]
        for attempt in range(1, 4):
            try:
                self._ws.append_rows(rows, value_input_option="USER_ENTERED",
                                     insert_data_option="INSERT_ROWS")
                self._sent += len(rows)
                log.info(f"  🟢 Sheets: +{len(rows)} rows (session {self._sent})")
                return True
            except Exception as e:
                err = str(e).lower()
                if any(k in err for k in ("429", "quota", "rate")):
                    time.sleep(30 * attempt)
                elif "401" in err or "403" in err:
                    log.error(f"  Sheets auth error: {e}"); self._ok = False; return False
                else:
                    time.sleep(6 * attempt)
        log.error("  Sheets: 3 attempts failed — batch skipped.")
        return False

# ═══════════════════════════════════════════════════════════════════
#   💾  STATE
# ═══════════════════════════════════════════════════════════════════

def hash_key(s: str) -> str:
    """SHA256 of a normalised string. Used to store emails/domains in the
    committed data/*.json as one-way hashes (repo is PUBLIC) while keeping
    dedup working: hash(email) collides iff the emails are equal.
    Real emails/domains are still written to the private Google Sheet + Excel."""
    return hashlib.sha256(str(s).lower().strip().encode()).hexdigest()


def _default_state() -> dict:
    return {"version": "4.0", "query_counter": 0, "used_query_hashes": [],
            "emails": {}, "seen_domains": [], "session_count": 0,
            "created_at": datetime.now().isoformat(),
            "last_updated": datetime.now().isoformat()}


def load_state() -> dict:
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                s = json.load(f)
            for k, v in _default_state().items():
                s.setdefault(k, v)
            log.info(f"📂 State: {len(s['emails'])} contacts, "
                     f"query #{s['query_counter']}, {len(s['seen_domains'])} domains")
            return s
        except Exception as e:
            log.warning(f"State load error ({e}). Fresh start.")
    return _default_state()


def save_state(s: dict):
    s["last_updated"] = datetime.now().isoformat()
    if len(s["used_query_hashes"]) > 60000:
        s["used_query_hashes"] = s["used_query_hashes"][-50000:]
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(s, f, indent=2, ensure_ascii=False)


def load_visited() -> set:
    if os.path.exists(VISITED_FILE):
        try:
            with open(VISITED_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def save_visited(v: set):
    with open(VISITED_FILE, "w", encoding="utf-8") as f:
        json.dump(list(v)[-120000:], f, ensure_ascii=False)


def reset_state():
    for p in [STATE_FILE, VISITED_FILE]:
        if os.path.exists(p):
            os.remove(p)
    log.info("🗑️  State cleared.")

# ═══════════════════════════════════════════════════════════════════
#   🚀  HARVESTER
# ═══════════════════════════════════════════════════════════════════

class EmailHarvester:

    def __init__(self, anthropic_key: str = ""):
        self.state        = load_state()
        self.visited      = load_visited()
        self.seen_domains = set(self.state["seen_domains"])
        self.used_hashes  = set(self.state["used_query_hashes"])

        self.google    = GoogleClient(API_PAIRS)
        self.http      = requests.Session()
        self.http.max_redirects = 5
        self.mx        = MXResolver()
        self.classifier= CompanyValidator(anthropic_key)
        self.verifier  = ContactVerifier(self.mx)
        self.mailbox   = MailboxVerifier(MAILBOX_PROVIDER, MAILBOX_API_KEYS)
        self.qgen      = QueryGenerator(self.state["query_counter"], self.used_hashes)

        self.running   = True
        self._buf      = []
        self._warned_nocredits = False

        try:                       # signal handlers only work in the main thread
            signal.signal(signal.SIGINT,  self._stop)
            signal.signal(signal.SIGTERM, self._stop)
        except (ValueError, RuntimeError):
            pass                   # running inside a web worker thread — skip
        init_excel(OUTPUT_EXCEL)

        self.sheets = SheetsUploader(SHEETS_SPREADSHEET_ID,
                                     SHEETS_WORKSHEET_GID, SHEETS_CREDS_FILE)
        self.sheets.connect()

    def _stop(self, sig, frame):
        print()
        log.info("⏸️  Stop — finishing current URL, then saving...")
        self.running = False

    def _flush(self):
        if self._buf:
            append_to_excel(OUTPUT_EXCEL, self._buf)
            self.sheets.append_rows(self._buf)
            self._buf = []
        self.state["query_counter"]     = self.qgen.counter
        self.state["used_query_hashes"] = list(self.used_hashes)
        self.state["seen_domains"]      = list(self.seen_domains)
        save_state(self.state)
        save_visited(self.visited)

    def _process_url(self, url: str) -> bool:
        if not url or not is_company_url(url):
            return False
        domain = registrable_domain(url)
        dh = hash_key(domain) if domain else ""
        if not domain or dh in self.seen_domains:
            return False
        self.seen_domains.add(dh)             # one shot per domain per DB (stored hashed)

        log.info(f"  🌐 {domain}")
        data = scrape_company(url, self.http, self.visited)

        if _looks_like_junk(data["company"]):
            data["company"] = domain         # never store listicle titles

        email = data["email"]
        eh = hash_key(email) if email else ""
        if not email or eh in self.state["emails"]:
            return False

        # ── Deliverability + liveness gauntlet ──────────────────
        status, score, reason = self.verifier.verify(
            email, data["domain"], data["page_text"], data["alive"])
        if status == "dead" or score < MIN_SCORE:
            log.info(f"    ❌ {status}/{score} [{reason}]  {email}")
            return False

        # ── Right KIND of company? (free, before paid API) ──────
        ok_kind, method = self.classifier.validate(
            data["company"], data["url"], data["page_text"])
        if not ok_kind:
            log.info(f"    ❌ wrong-niche [{method}]  {email} → {data['company']}")
            return False

        # ── Real mailbox check (rotating API keys — runs LAST) ──
        # When credits run out we DON'T stop: we fall back to the free
        # gauntlet and tag the row 'unverified'. Same 5 values → same
        # columns, so the Google Sheet layout never changes.
        no_credits = self.mailbox.enabled() and self.mailbox.exhausted()
        if no_credits and not self._warned_nocredits:
            log.warning("  ✉️  Mailbox credits exhausted — continuing on FREE gauntlet; "
                        "new rows tagged 'unverified'. Top up to resume mailbox checks.")
            self._warned_nocredits = True

        mbox = self.mailbox.verify(email)   # 'off' if disabled; 'unknown' if no keys left

        # Quality drops apply ONLY while we still have credits to verify.
        if not no_credits:
            if mbox == "undeliverable":
                log.info(f"    ❌ mailbox undeliverable  {email}")
                return False
            if mbox == "catch_all" and MAILBOX_DROP_CATCHALL:
                log.info(f"    ❌ catch-all dropped  {email}")
                return False
            if mbox == "unknown" and MAILBOX_DROP_UNKNOWN and self.mailbox.enabled():
                log.info(f"    ❌ mailbox unverifiable dropped  {email}")
                return False

        if mbox == "deliverable":
            score = min(score + 15, 100)

        # Deliverability marker for the 'Validated By' column (no new columns).
        if not self.mailbox.enabled():
            mtag = ""                       # mailbox check disabled
        elif no_credits or mbox == "unknown":
            mtag = "+unverified"            # collected on free gauntlet only
        else:
            mtag = f"+{mbox}"               # +deliverable / +catch_all

        log.info(f"    ✅ {status}/{score}{mtag}  {email}  →  {data['company']}")
        rec = {"company": data["company"], "url": data["url"],
               "status": status, "score": score,
               "method": f"{method}+mx{mtag}",
               "date": datetime.now().strftime("%Y-%m-%d %H:%M")}
        # Committed state (public repo): key by SHA256(email), value carries NO
        # raw email/domain/url/company — only non-identifying metadata for stats.
        self.state["emails"][eh] = {"status": status, "score": score,
                                    "date": rec["date"]}
        self._buf.append({"email": email, **rec})   # real email → private Excel/Sheets
        if len(self._buf) >= 12:
            self._flush()
        return True

    def _google_phase(self) -> int:
        log.info("─" * 64)
        log.info(f"🔍  GOOGLE PHASE — {MAX_GOOGLE_QUERIES} queries")
        queries = self.qgen.next_batch(MAX_GOOGLE_QUERIES)
        log.info(f"  Generated {len(queries)} queries (counter #{self.qgen.counter})")
        added = 0
        for i, (q, lang) in enumerate(queries, 1):
            if not self.running:
                break
            if self.google.exhausted():
                log.warning("  ⛔ Google quota exhausted for today — ending "
                            "search phase early (no more keys).")
                break
            log.info(f"\n  [G {i}/{len(queries)}] ({lang}) {q}")
            urls = []
            for page in range(SEARCH_PAGES):
                batch = self.google.search(q, start=1 + page * RESULTS_PER_QUERY,
                                           lang=lang)
                urls += batch
                if len(batch) < RESULTS_PER_QUERY:
                    break  # выдача кончилась, вторую страницу не тратим
            log.info(f"    → {len(urls)} URLs ({SEARCH_PAGES} pages max)")
            for url in urls:
                if not self.running:
                    break
                if self._process_url(url):
                    added += 1
                time.sleep(DELAY_URL)
            time.sleep(DELAY_QUERY)
        return added

    def run(self):
        self.state["session_count"] += 1
        before = len(self.state["emails"])
        log.info("=" * 64)
        log.info(f"  IT Email Harvester v4.0  │  Session #{self.state['session_count']}")
        log.info(f"  Profile: {PROFILE['name']}   MIN_SCORE={MIN_SCORE}")
        log.info(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        log.info(f"  Contacts in DB    : {before}")
        log.info(f"  Domains processed : {len(self.seen_domains)}")
        log.info(f"  Claude validation : {'ON ✅' if self.classifier.claude.enabled() else 'OFF (rules only)'}")
        log.info(f"  Mailbox API       : {('ON ✅ ('+self.mailbox.provider+')') if self.mailbox.enabled() else 'OFF (free gauntlet only)'}")
        log.info("=" * 64)

        g_new = self._google_phase()
        self._flush()

        after = len(self.state["emails"])
        v, vf = self.classifier.stats, self.verifier.stats
        log.info("=" * 64)
        log.info(f"  Session complete  │  +{after - before} new   total {after}")
        log.info(f"  ── Deliverability gauntlet ──")
        log.info(f"    live kept        : {vf['live']}")
        log.info(f"    risky (dropped)  : {vf['risky']}")
        log.info(f"    rejected no-MX   : {vf['no_mx']}")
        log.info(f"    rejected mismatch: {vf['mismatch']}")
        log.info(f"    rejected bad-fmt : {vf['bad_format']}")
        log.info(f"    rejected dead-sit: {vf['dead_site']}")
        log.info(f"  ── Niche classifier ──")
        log.info(f"    accepted         : {v['accepted']}")
        log.info(f"    rejected         : {v['rejected']}")
        log.info(f"    uncertain dropped: {v['uncertain_dropped']}")
        log.info(f"    claude calls     : {v['claude_calls']}")
        if self.mailbox.enabled():
            mb = self.mailbox.stats
            log.info(f"  ── Mailbox API ({self.mailbox.provider}) ──")
            log.info(f"    deliverable      : {mb['deliverable']}")
            log.info(f"    catch-all        : {mb['catch_all']}")
            log.info(f"    undeliverable    : {mb['undeliverable']}")
            log.info(f"    unknown          : {mb['unknown']}")
            log.info(f"    API credits used : {mb['calls']}")
            dist = ", ".join(f"#{i+1}:{c}" + ("✗" if i in self.mailbox.dead else "")
                             for i, c in enumerate(self.mailbox.per_key))
            log.info(f"    per-key calls    : {dist}")
        log.info(f"  Google API calls  : {self.google.calls}")
        log.info(f"  MX cache size     : {len(self.mx.cache)}")
        log.info(f"  Output (local)    : {OUTPUT_EXCEL}")
        log.info(f"  Google Sheets rows: {self.sheets._sent}")
        log.info("=" * 64)

# ═══════════════════════════════════════════════════════════════════
#   🔧  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

def run_once() -> dict:
    """Single batch run — called by the web service. Returns a summary dict."""
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    h = EmailHarvester(anthropic_key=key)
    before = len(h.state["emails"])
    h.run()
    after = len(h.state["emails"])
    return {
        "parser": "b2b",
        "added": after - before,
        "total_in_db": after,
        "sheet_rows_this_run": getattr(h.sheets, "_sent", 0),
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="IT Company Email Harvester v4.0 — live companies, deliverable contacts")
    parser.add_argument("--reset", action="store_true", help="clear state, start fresh")
    parser.add_argument("--key", type=str, default="", help="Claude API key (optional)")
    parser.add_argument("--min-score", type=int, default=None, help="quality bar 0-100")
    parser.add_argument("--max-queries", type=int, default=None, help="queries this session")
    parser.add_argument("--mailbox-provider", type=str, default=None,
                        help="millionverifier | reoon (enables real mailbox check)")
    parser.add_argument("--mailbox-keys", type=str, default=None,
                        help="comma-separated API keys (rotated round-robin)")
    args = parser.parse_args()

    if args.min_score is not None:
        MIN_SCORE = args.min_score
    if args.max_queries is not None:
        MAX_GOOGLE_QUERIES = args.max_queries
    if args.mailbox_provider is not None:
        MAILBOX_PROVIDER = args.mailbox_provider.strip()
    if args.mailbox_keys is not None:
        MAILBOX_API_KEYS = [k.strip() for k in args.mailbox_keys.split(",") if k.strip()]
    if args.reset:
        reset_state()

    key = args.key.strip() or ANTHROPIC_API_KEY.strip()
    EmailHarvester(anthropic_key=key).run()
